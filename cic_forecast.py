#!/usr/bin/env python3
"""
CIC Forecasting — Bank of Thailand — single-file pipeline

*** PRIMARY KPI: 1-month-ahead END-OF-MONTH (EOM) CIC level RMSE. ***
Daily ΔCIC RMSE (below) is a secondary, no-regression guardrail — the desk
does not get scored on daily accuracy, it gets scored on whether the EOM
level forecast made at each month-end is close to the level that actually
materialises 1 month later. Everything downstream (model selection, the
blend, Gate 0) is judged on the EOM number, not the daily one.

See CIC_model_document.md for the full problem statement, methodology and
results — this docstring is a map of what's in this file, not the analysis.

Two things live in one script now (merged 2026-08, previously
cic_forecast.py + cic_forecast_v2.py):

  (1) The v1 daily pipeline — data loading, the daily models, GARCH,
      residual diagnostics, and the user-facing Excel/figure exports:
        Daily_Baseline          the 2022 EViews model — OLS on 55 calendar
                                 dummies + ARIMA(1,0,1) on the residuals
        Daily_AdaptiveDrift     + local-level state-space drift replacing
                                 the constant
        Daily_AdaptiveSeasonal  + trailing-window OLS betas and a
                                 smooth-trend state
      Estimation strategy: two-step ARIMAX (OLS for the mean equation, then
      ARIMA(1,0,1) on residuals) — numerically equivalent to joint SARIMAX-MLE
      in large samples (Frisch-Waugh theorem) but converges in seconds rather
      than minutes, making rolling backtests practical.

  (2) The v2 EOM-level candidate harness — expanding-window backtest scored
      on the primary KPI, plus the monthly and blended candidates:
        Monthly_SARIMA, Monthly_UC, Daily_LevelTrend, Blend_*
      selection/holdout protocol, DM significance test, forecast combination.

Bug fix (2026-08, §3(a) of CIC_model_document.md): generate_future_exog()
used to build its daily calendar with pd.bdate_range() alone, which includes
Thai public holidays that input.xlsx's RAW sheet (and BOT's actual trading
calendar) never contains. Every daily model was summing its forecast over
"phantom" non-trading days, each adding a spurious drift-plus-dummy
contribution to the EOM total — worst in Songkran-heavy April, matching the
worst seasonal bias in the diagnostics. Fixed in generate_future_exog();
see its docstring for the two calling conventions.

Usage:
  python cic_forecast.py                                  # full v1 diagnostic pipeline (default; unchanged CLI)
  python cic_forecast.py --eom                             # v2 EOM-level KPI harness, all candidates
  python cic_forecast.py --eom --gate0                     # EOM harness: post-fix reproduction/sanity check only
  python cic_forecast.py --eom --models baseline,monthly_sarima,level_trend

Literature:
  Box & Jenkins (1970), Engle (1982), Bollerslev (1986),
  Bai & Perron (1998), Anderson & Gascon (2009), Tashman (2000),
  Diebold & Mariano (1995), Harvey/Leybourne/Newbold (1997), Timmermann (2006)
"""

import warnings
warnings.filterwarnings('ignore')

import argparse
import os
import pickle
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.cm as cm
import matplotlib.patches as mpatches
from scipy import stats

from sklearn.linear_model import LinearRegression
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.stattools import adfuller, acf
from statsmodels.stats.diagnostic import het_arch, acorr_ljungbox
from statsmodels.tsa.statespace.structural import UnobservedComponents
from statsmodels.tsa.statespace.sarimax import SARIMAX
from arch import arch_model

np.random.seed(42)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def _build_sk_ny_dummies(df, hol):
    """
    Build Songkran (SK) and New Year (NY) pre/post holiday dummies.

    Holiday sheet only covers 2014-2026.  To ensure stable OLS coefficient
    estimates (one event per year across all training years), SK and NY dates
    are extended back to 1997 using fixed-calendar heuristics:
      SK  → April 13–15 (Thai calendar, fixed)
      NY  → January 1 + December 31

    All four dummies are mutually exclusive (no overlapping coverage):
      D_SK_PRE1  : last trading day before each Songkran block
      D_SK_POST1 : first trading day after each Songkran block
      D_NY_PRE1  : last trading day before each New Year block
      D_NY_POST1 : first trading day after each New Year block
    """
    trading_dates = sorted(df.index.normalize().tolist())

    sk_from_sheet = set(
        hol.loc[hol['Description'].str.contains('Songkran', case=False, na=False),
                'Date'].dt.normalize()
    )
    min_sk_year = min(d.year for d in sk_from_sheet) if sk_from_sheet else 2014
    sk_dates = set(sk_from_sheet)
    start_yr = df.index.year.min()
    for yr in range(start_yr, min_sk_year):
        for mday in [(4, 13), (4, 14), (4, 15)]:
            sk_dates.add(pd.Timestamp(yr, *mday))

    ny_from_sheet = set(
        hol.loc[hol['Description'].str.contains('New Year', case=False, na=False),
                'Date'].dt.normalize()
    )
    min_ny_year = min(d.year for d in ny_from_sheet) if ny_from_sheet else 2014
    ny_dates = set(ny_from_sheet)
    for yr in range(start_yr, min_ny_year):
        ny_dates.add(pd.Timestamp(yr, 1, 1))
        ny_dates.add(pd.Timestamp(yr, 12, 31))

    def holiday_blocks(hol_set):
        if not hol_set:
            return []
        sorted_h = sorted(hol_set)
        blocks, bs, be = [], sorted_h[0], sorted_h[0]
        for d in sorted_h[1:]:
            if (d - be).days <= 3:
                be = d
            else:
                blocks.append((bs, be))
                bs = be = d
        blocks.append((bs, be))
        return blocks

    def nearest_td(blocks, n_pre, n_post):
        pre_idx, post_idx = set(), set()
        for bs, be in blocks:
            pre_td  = [t for t in trading_dates if t < bs]
            post_td = [t for t in trading_dates if t > be]
            for lag in range(1, n_pre + 1):
                if lag <= len(pre_td):
                    pre_idx.add(pre_td[-lag])
            for lag in range(1, n_post + 1):
                if lag <= len(post_td):
                    post_idx.add(post_td[lag - 1])
        return pre_idx, post_idx

    sk_blocks = holiday_blocks(sk_dates)
    ny_blocks = holiday_blocks(ny_dates)

    sk_pre1_idx, sk_post1_idx = nearest_td(sk_blocks, 1, 1)
    ny_pre1_idx, ny_post1_idx = nearest_td(ny_blocks, 1, 1)

    idx = df.index.normalize()
    df['D_SK_PRE1']  = idx.isin(sk_pre1_idx).astype(float)
    df['D_SK_POST1'] = idx.isin(sk_post1_idx).astype(float)
    df['D_NY_PRE1']  = idx.isin(ny_pre1_idx).astype(float)
    df['D_NY_POST1'] = idx.isin(ny_post1_idx).astype(float)
    return df


def load_data(filepath='input.xlsx'):
    raw = pd.read_excel(filepath, sheet_name='RAW', header=1)
    raw['Date']     = pd.to_datetime(raw['Date'],     errors='coerce')
    raw['Currency'] = pd.to_numeric(raw['Currency'],  errors='coerce')
    raw = raw.sort_values('Date').reset_index(drop=True)
    raw['Change'] = raw['Currency'].diff()
    df = raw.dropna(subset=['Date', 'Change']).copy()
    dummy_cols = [c for c in df.columns if c.startswith('D_') or c.startswith('Date_')]
    for c in dummy_cols:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).astype(float)
    df = df.set_index('Date')
    df.index = df.index.normalize()

    df['D_PostCovid'] = (df.index >= pd.Timestamp('2020-04-01')).astype(float)
    t = np.arange(len(df), dtype=float)
    P_ann = 261.0
    for k in range(1, 4):
        df[f'sin_ann_{k}'] = np.sin(2 * np.pi * k * t / P_ann)
        df[f'cos_ann_{k}'] = np.cos(2 * np.pi * k * t / P_ann)

    hol = pd.read_excel(filepath, sheet_name='holiday')
    hol['Date'] = pd.to_datetime(hol['Date'])
    df = _build_sk_ny_dummies(df, hol)
    return df


def load_holiday(filepath='input.xlsx'):
    hol = pd.read_excel(filepath, sheet_name='holiday')
    hol['Date'] = pd.to_datetime(hol['Date'])
    return hol


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — REGRESSOR SETS
# ─────────────────────────────────────────────────────────────────────────────

DOM_COLS = [f'Date_{str(i).zfill(2)}' for i in range(2, 32)]
DOW_COLS = ['D_TUE', 'D_WED', 'D_THU', 'D_FRI']
WOM_COLS = ['D_WEEK2', 'D_WEEK3', 'D_WEEK4', 'D_WEEK5']
MON_COLS = ['D_JAN', 'D_FEB', 'D_MAR', 'D_APR', 'D_MAY', 'D_JUN',
            'D_JUL', 'D_AUG', 'D_SEP', 'D_OCT', 'D_NOV']
HOL_OLD  = ['D_PRE_LH1', 'D_PRE_LH3', 'D_POST_LH3', 'D_PRE_SH1', 'D_Covid_1st', 'D_LWD']
HOL_EXT  = ['D_SK_PRE1', 'D_SK_POST1', 'D_NY_PRE1', 'D_NY_POST1']
REGIME   = ['D_PostCovid']
FOURIER  = [f'{fn}_ann_{k}' for fn in ['sin', 'cos'] for k in range(1, 4)]

REGS = {
    'Daily_Baseline': DOM_COLS + DOW_COLS + WOM_COLS + MON_COLS + HOL_OLD,
}

# Display labels == model keys, so figures, Excel headers and the documentation
# all use one vocabulary (<Frequency>_<Method>).
BASE_LABELS = {
    'Daily_Baseline':         'Daily_Baseline',
    'Daily_AdaptiveDrift':    'Daily_AdaptiveDrift',
    'Daily_AdaptiveSeasonal': 'Daily_AdaptiveSeasonal',
}

COLORS = {
    'Daily_Baseline':         '#d62728',
    'Daily_AdaptiveDrift':    '#9467bd',
    'Daily_AdaptiveSeasonal': '#2ca02c',
}


def model_label(mname, train_label):
    return f'{BASE_LABELS.get(mname, mname)} ({train_label})'


def get_X(df, model_name):
    cols = [c for c in REGS[model_name] if c in df.columns]
    return df[cols].astype(float).values, cols


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — FUTURE EXOG GENERATION (for seasonal forecast dot)
# ─────────────────────────────────────────────────────────────────────────────

def _holiday_blocks_from_set(hol_set):
    if not hol_set:
        return []
    sorted_h = sorted(hol_set)
    blocks, bs, be = [], sorted_h[0], sorted_h[0]
    for d in sorted_h[1:]:
        if (d - be).days <= 3:
            be = d
        else:
            blocks.append((bs, be))
            bs = be = d
    blocks.append((bs, be))
    return blocks


def generate_future_exog(model_name, start_str, end_str, hol, actual_dates=None):
    """
    Build the exogenous variable matrix for trading dates in [start, end].
    Used to forecast CIC for months beyond the last data point.

    Bug fix (2026-08, §3(a) of CIC_model_document.md): input.xlsx's RAW sheet
    only contains days BOT actually trades (business days minus Thai public
    holidays — Songkran, Labour Day, etc.), but this function used to build
    the matrix from pd.bdate_range() alone, which includes those holidays.
    Every daily model then summed its forecast over "phantom" days that don't
    exist in the actual series, each adding a spurious intercept-drift-plus-
    dummy contribution to the EOM sum. April (Songkran) was hit hardest,
    matching the worst seasonal bias in the diagnostics.

    Two calling conventions now:
      actual_dates given  (backtesting a historical month) — restrict the
        matrix to exactly those realised trading days (df_next.index). This
        is always correct since the actual calendar for that month is known.
      actual_dates=None   (forecasting a genuinely future, not-yet-realised
        month) — fall back to business days minus the holiday sheet's dates,
        which is the best available estimate of the future trading calendar.
    """
    if actual_dates is not None:
        dates = pd.DatetimeIndex(sorted({pd.Timestamp(d).normalize() for d in actual_dates}))
    else:
        bdays = pd.bdate_range(start_str, end_str)
        hol_dates = set(hol['Date'].dt.normalize())
        dates = pd.DatetimeIndex([d for d in bdays if d.normalize() not in hol_dates])
    if len(dates) == 0:
        return pd.DataFrame()

    fut = pd.DataFrame(index=dates)
    fut.index = fut.index.normalize()

    # Day of month
    for d in range(1, 32):
        fut[f'Date_{d:02d}'] = (fut.index.day == d).astype(float)

    # Day of week
    for i, nm in enumerate(['D_MON', 'D_TUE', 'D_WED', 'D_THU', 'D_FRI']):
        fut[nm] = (fut.index.dayofweek == i).astype(float)

    # Month
    mnames = ['D_JAN','D_FEB','D_MAR','D_APR','D_MAY','D_JUN',
              'D_JUL','D_AUG','D_SEP','D_OCT','D_NOV','D_DEC']
    for m, nm in enumerate(mnames, 1):
        fut[nm] = (fut.index.month == m).astype(float)

    # Week of month (1-5 based on day)
    for w in range(1, 6):
        fut[f'D_WEEK{w}'] = ((fut.index.day - 1) // 7 + 1 == w).astype(float)

    # Last working day of month
    lwd = set()
    for (yr, mo), _ in fut.groupby([fut.index.year, fut.index.month]):
        subset = fut[(fut.index.year == yr) & (fut.index.month == mo)]
        lwd.add(subset.index[-1])
    fut['D_LWD'] = fut.index.isin(lwd).astype(float)

    # COVID dummy: 0 for future
    fut['D_Covid_1st'] = 0.0
    fut['D_PostCovid']  = 1.0  # post-April-2020

    # Fourier terms (relative to start of series ~1997)
    origin = pd.Timestamp('1997-08-29')
    P_ann  = 261.0
    t_vals = np.array([(d - origin).days * 5 / 7 for d in fut.index], dtype=float)
    for k in range(1, 4):
        fut[f'sin_ann_{k}'] = np.sin(2 * np.pi * k * t_vals / P_ann)
        fut[f'cos_ann_{k}'] = np.cos(2 * np.pi * k * t_vals / P_ann)

    # Holiday-based dummies from holiday sheet
    fut_dates_list = sorted(fut.index.tolist())
    window_start = pd.Timestamp(start_str) - pd.Timedelta(days=14)
    window_end   = pd.Timestamp(end_str)   + pd.Timedelta(days=14)
    near_hol = set(
        hol.loc[(hol['Date'] >= window_start) & (hol['Date'] <= window_end), 'Date'].dt.normalize()
    )
    blocks    = _holiday_blocks_from_set(near_hol)
    long_blk  = [(bs, be) for bs, be in blocks if (be - bs).days >= 2]
    short_blk = [(bs, be) for bs, be in blocks if (be - bs).days < 2]

    pre_lh3, pre_lh1, post_lh3, pre_sh1 = set(), set(), set(), set()
    for bs, be in long_blk:
        pre_td  = [t for t in fut_dates_list if t < bs]
        post_td = [t for t in fut_dates_list if t > be]
        for lag in range(1, 4):
            if lag <= len(pre_td):
                pre_lh3.add(pre_td[-lag])
        if pre_td:
            pre_lh1.add(pre_td[-1])
        for lag in range(1, 4):
            if lag <= len(post_td):
                post_lh3.add(post_td[lag - 1])
    for bs, be in short_blk:
        pre_td = [t for t in fut_dates_list if t < bs]
        if pre_td:
            pre_sh1.add(pre_td[-1])

    fi = fut.index
    fut['D_PRE_LH1']  = fi.isin(pre_lh1).astype(float)
    fut['D_PRE_LH3']  = fi.isin(pre_lh3).astype(float)
    fut['D_POST_LH3'] = fi.isin(post_lh3).astype(float)
    fut['D_PRE_SH1']  = fi.isin(pre_sh1).astype(float)

    # SK/NY dummies for future window (Songkran = April 13-15, NY = Jan 1 / Dec 31)
    fut_yr_range = range(fut.index.year.min(), fut.index.year.max() + 1)
    sk_fut = set()
    ny_fut = set()
    sk_from_sheet = set(
        hol.loc[hol['Description'].str.contains('Songkran', case=False, na=False),
                'Date'].dt.normalize()
    )
    ny_from_sheet = set(
        hol.loc[hol['Description'].str.contains('New Year', case=False, na=False),
                'Date'].dt.normalize()
    )
    for yr in fut_yr_range:
        # Use sheet if available, else heuristic
        yr_sk = {d for d in sk_from_sheet if d.year == yr}
        if yr_sk:
            sk_fut |= yr_sk
        else:
            for md in [(4, 13), (4, 14), (4, 15)]:
                sk_fut.add(pd.Timestamp(yr, *md))
        yr_ny = {d for d in ny_from_sheet if d.year == yr}
        if yr_ny:
            ny_fut |= yr_ny
        else:
            ny_fut.add(pd.Timestamp(yr, 1, 1))
            ny_fut.add(pd.Timestamp(yr, 12, 31))

    sk_blk = _holiday_blocks_from_set(sk_fut)
    ny_blk = _holiday_blocks_from_set(ny_fut)

    sk_pre1, sk_post1 = set(), set()
    ny_pre1, ny_post1 = set(), set()
    for bs, be in sk_blk:
        pre_td  = [t for t in fut_dates_list if t < bs]
        post_td = [t for t in fut_dates_list if t > be]
        if pre_td:  sk_pre1.add(pre_td[-1])
        if post_td: sk_post1.add(post_td[0])
    for bs, be in ny_blk:
        pre_td  = [t for t in fut_dates_list if t < bs]
        post_td = [t for t in fut_dates_list if t > be]
        if pre_td:  ny_pre1.add(pre_td[-1])
        if post_td: ny_post1.add(post_td[0])

    fut['D_SK_PRE1']  = fi.isin(sk_pre1).astype(float)
    fut['D_SK_POST1'] = fi.isin(sk_post1).astype(float)
    fut['D_NY_PRE1']  = fi.isin(ny_pre1).astype(float)
    fut['D_NY_POST1'] = fi.isin(ny_post1).astype(float)

    cols = [c for c in REGS[model_name] if c in fut.columns]
    return fut[cols]


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — TWO-STEP ARIMAX
# ─────────────────────────────────────────────────────────────────────────────

class TwoStepARIMAX:
    """
    Two-step ARIMAX:
      Step 1) OLS on calendar/dummy regressors  →  beta, residuals
      Step 2) ARIMA(1,0,1) on OLS residuals     →  phi, theta, sigma2
    """
    def __init__(self):
        self.ols = self.arima = self.resid = self.fitted = None
        self.n_obs = self.n_params = None

    def fit(self, y, X):
        y, X = np.asarray(y, float), np.asarray(X, float)
        n, p = X.shape
        self.ols   = LinearRegression(fit_intercept=True).fit(X, y)
        ols_fit    = self.ols.predict(X)
        ols_res    = y - ols_fit
        self.arima = ARIMA(ols_res, order=(1, 0, 1), trend='n').fit(
            method='innovations_mle')
        self.resid  = self.arima.resid
        self.fitted = ols_fit + self.arima.fittedvalues
        self.n_obs  = n
        k = (p + 1) + 2 + 1
        self.n_params = k
        self._logL = self.arima.llf
        self.aic   = -2 * self._logL + 2 * k
        self.bic   = -2 * self._logL + k * np.log(n)
        pn = self.arima.param_names
        pv = self.arima.params
        ps = pd.Series(pv, index=pn)
        self.ar1   = float(ps.get('ar.L1',   np.nan))
        self.ma1   = float(ps.get('ma.L1',   np.nan))
        self.sigma = float(np.sqrt(ps.get('sigma2', np.nan)))
        return self

    def forecast(self, X_future):
        X_future = np.asarray(X_future, float)
        mean_fc  = self.ols.predict(X_future)
        arima_fc = self.arima.forecast(steps=len(X_future))
        return mean_fc + np.asarray(arima_fc)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4b — STATE-SPACE TREND MODEL (Model D)
# ─────────────────────────────────────────────────────────────────────────────

class AdaptiveDriftModel:
    """
    Model D: adaptive-drift forecaster — two-step state-space.

    Step 1: OLS on the same calendar dummy matrix as Daily_Baseline (concentrates out
            regression betas analytically, same as TwoStepARIMAX Step 1).
    Step 2: UnobservedComponents on the OLS residuals — no exog, so MLE only
            optimises 2–3 variance parameters (20–50× faster than joint UC).

    Daily_AdaptiveDrift  — local level on ΔCIC residuals:
            resid_t = ν_t + u_t,  ν_t = ν_{t-1} + ζ_t  (AR(1) irregular)
          ν_t is the stochastic drift; adapts as the regime changes.
          EOM forecast = L_last + Σ (OLS_mean + UC_level_forecast).

    D2_smooth — smooth trend on CIC-level residuals:
            resid_t = ℓ_t + b_t·t + u_t,  b_t drifts slowly (level var = 0).
          Directly targets the end-of-month level KPI.
          EOM forecast = terminal UC level forecast + OLS mean.
    """

    VARIANTS = {
        'Daily_AdaptiveDrift': {'endog': 'change', 'level': 'local level'},
    }

    def __init__(self, variant='Daily_AdaptiveDrift'):
        if variant not in self.VARIANTS:
            raise ValueError(f'Unknown variant {variant!r}. Choose from {list(self.VARIANTS)}')
        self.variant = variant
        self._cfg    = self.VARIANTS[variant]
        self.ols     = None
        self.uc_res  = None
        self.fitted  = None
        self.resid   = None
        self.aic     = self.bic = np.nan

    def fit(self, y_change, X, y_level=None):
        # Step 1 — OLS (concentrates calendar betas out of Kalman MLE)
        self.ols    = LinearRegression(fit_intercept=True).fit(X, y_change)
        ols_resid   = y_change - self.ols.predict(X)
        # Step 2 — local-level UC on residuals (2 variance params only → fast)
        mod         = UnobservedComponents(endog=ols_resid, level='local level', autoregressive=1)
        self.uc_res = mod.fit(disp=False, method='bfgs', maxiter=300)
        self.fitted = self.ols.predict(X) + np.asarray(self.uc_res.fittedvalues)
        self.resid  = y_change - self.fitted
        self.aic    = self.uc_res.aic
        self.bic    = self.uc_res.bic
        return self

    def forecast(self, X_future):
        ols_fc = self.ols.predict(X_future)
        uc_fc  = self.uc_res.get_forecast(steps=len(X_future))
        return ols_fc + np.asarray(uc_fc.predicted_mean)

    def smoothed_drift(self):
        """Smoothed local-level state ν_t on ΔCIC residuals (= adaptive drift)."""
        return self.uc_res.smoother_results.smoothed_state[0]


class AdaptiveSeasonalModel:
    """
    Daily_AdaptiveSeasonal: smooth-trend adaptive forecaster — two-step state-space.

    Step 1: OLS on a TRAILING WINDOW of the last TRAILING_MONTHS calendar months
            (default 60 = 5 years) using the same 55-dummy matrix as Daily_Baseline.
            This lets seasonal betas adapt to recent patterns.
    Step 2: UnobservedComponents with level='smooth trend' + autoregressive=1
            on the OLS residuals from the full history (using trailing-window betas).
            Smooth trend: level variance fixed to 0, slope drifts (better for
            slow-moving level forecasts; targets EOM level KPI).
    """
    TRAILING_MONTHS = 60  # tunable: ~5 years of recent data for OLS step

    def __init__(self):
        self.ols = None
        self.uc_res = None
        self.fitted = None
        self.resid = None
        self.aic = self.bic = np.nan

    def fit(self, y_change, X, dates=None):
        """
        y_change : full-history array of daily ΔCIC
        X        : full-history dummy matrix (same 55 cols as Daily_Baseline)
        dates    : optional pd.DatetimeIndex aligned with y_change/X;
                   used to select trailing window. If None, uses last
                   TRAILING_MONTHS*22 rows as proxy.
        """
        y_change = np.asarray(y_change, float)
        X = np.asarray(X, float)
        n = len(y_change)

        # --- Step 1: trailing-window OLS ---
        if dates is not None:
            cutoff = dates[-1] - pd.DateOffset(months=self.TRAILING_MONTHS)
            mask = dates >= cutoff
        else:
            # fallback: approximate 22 trading days/month
            tw = min(self.TRAILING_MONTHS * 22, n)
            mask = np.zeros(n, bool)
            mask[-tw:] = True

        X_tw = X[mask]
        y_tw = y_change[mask]
        self.ols = LinearRegression(fit_intercept=True).fit(X_tw, y_tw)

        # --- Compute OLS residuals on FULL history using trailing-window betas ---
        ols_fit_full = self.ols.predict(X)
        ols_resid_full = y_change - ols_fit_full

        # --- Step 2: UC smooth trend on full residuals ---
        mod = UnobservedComponents(
            endog=ols_resid_full,
            level='smooth trend',
            autoregressive=1
        )
        self.uc_res = mod.fit(disp=False, method='bfgs', maxiter=300)
        self.fitted = ols_fit_full + np.asarray(self.uc_res.fittedvalues)
        self.resid = y_change - self.fitted
        self.aic = self.uc_res.aic
        self.bic = self.uc_res.bic
        return self

    def forecast(self, X_future):
        ols_fc = self.ols.predict(np.asarray(X_future, float))
        uc_fc = self.uc_res.get_forecast(steps=len(X_future))
        return ols_fc + np.asarray(uc_fc.predicted_mean)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 — METRICS AND DIAGNOSTICS
# ─────────────────────────────────────────────────────────────────────────────

def compute_metrics(actual, predicted):
    e = np.asarray(actual, float) - np.asarray(predicted, float)
    return {'RMSE': np.sqrt(np.mean(e**2)), 'MAE': np.mean(np.abs(e)),
            'ResidSD': np.std(e, ddof=1), 'Bias': np.mean(e),
            'n': len(e), 'errors': e}


def run_diagnostics(residuals, label=''):
    res = np.asarray(residuals, float)
    res = res[~np.isnan(res)]
    adf  = adfuller(res, autolag='AIC')
    arch = het_arch(res, nlags=10)
    lb   = acorr_ljungbox(res, lags=[10, 20], return_df=True)
    out  = {'adf_stat': adf[0], 'adf_pval': adf[1],
            'arch_stat': arch[0], 'arch_pval': arch[1],
            'lb_pval_10': float(lb['lb_pvalue'].iloc[0]),
            'lb_pval_20': float(lb['lb_pvalue'].iloc[1])}
    if label:
        print(f'\n  [{label}] Residual Diagnostics:')
        print(f'    ADF stationary:   stat={adf[0]:7.3f}  p={adf[1]:.4f}  '
              f'{"✓ stationary" if adf[1]<0.05 else "⚠ non-stationary"}')
        print(f'    ARCH-LM(10):      stat={arch[0]:7.3f}  p={arch[1]:.4f}  '
              f'{"⚠ ARCH effects → GARCH warranted" if arch[1]<0.05 else "✓ no ARCH"}')
        print(f'    Ljung-Box p(10/20): {out["lb_pval_10"]:.4f} / {out["lb_pval_20"]:.4f}  '
              f'{"⚠ autocorrelation" if out["lb_pval_10"]<0.05 else "✓ white noise"}')
    return out


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6 — BACKTESTING
# ─────────────────────────────────────────────────────────────────────────────

def rolling_backtest(df, model_names, windows):
    results = {m: {} for m in model_names}
    for train_end, eval_start, eval_end in windows:
        wlabel = f'{eval_start[:7]}→{eval_end[:7]}'
        df_tr  = df.loc[:train_end]
        df_ev  = df.loc[eval_start:eval_end]
        if len(df_tr) < 200 or len(df_ev) < 5:
            continue
        print(f'\n  Window: train≤{train_end}, eval {eval_start}→{eval_end} ({len(df_ev)} obs)')
        for mname in model_names:
            X_tr, _ = get_X(df_tr, mname)
            X_ev, _ = get_X(df_ev, mname)
            try:
                mdl  = TwoStepARIMAX().fit(df_tr['Change'].values, X_tr)
                pred = mdl.forecast(X_ev)
                m    = compute_metrics(df_ev['Change'].values, pred)
                lbl  = BASE_LABELS[mname]
                print(f'    {lbl:<20}  RMSE={m["RMSE"]:.3f}  MAE={m["MAE"]:.3f}  ResidSD={m["ResidSD"]:.3f}')
            except Exception as exc:
                print(f'    ⚠ {mname}: {exc}')
                m = {'RMSE': np.nan, 'MAE': np.nan, 'ResidSD': np.nan, 'n': 0}
            results[mname][wlabel] = m
    return results


def horizon_rmse_monthly(df, model_names, monthly_origins, horizons=(1, 5, 10, 22)):
    store = {m: {h: [] for h in horizons} for m in model_names}
    for origin_str in monthly_origins:
        origin   = pd.Timestamp(origin_str)
        df_train = df[df.index < origin]
        df_after = df[df.index >= origin]
        max_h    = max(horizons)
        if len(df_train) < 200 or len(df_after) < max_h:
            continue
        print(f'\n  Horizon refit at {origin_str} ({len(df_train)} train obs)...')
        for mname in model_names:
            X_tr, _ = get_X(df_train, mname)
            X_fc, _ = get_X(df_after.iloc[:max_h], mname)
            try:
                mdl     = TwoStepARIMAX().fit(df_train['Change'].values, X_tr)
                fc_seq  = mdl.forecast(X_fc)
                act_seq = df_after['Change'].iloc[:max_h].values
                for h in horizons:
                    if h <= len(fc_seq):
                        store[mname][h].append((fc_seq[h-1], act_seq[h-1]))
            except Exception as exc:
                print(f'    ⚠ {mname}: {exc}')
    h_rmse = {}
    for mname in model_names:
        h_rmse[mname] = {}
        for h in horizons:
            pairs = store[mname][h]
            if pairs:
                e = np.array([p - a for p, a in pairs])
                h_rmse[mname][h] = np.sqrt(np.mean(e**2))
            else:
                h_rmse[mname][h] = np.nan
    return h_rmse


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6b — MONTH-END LEVEL BACKTEST (Model D evaluation)
# ─────────────────────────────────────────────────────────────────────────────

def month_end_eom_backtest(df, hol, start_year=2020, end_year=2025):
    """
    Rolling monthly backtest: primary KPI = 1-month-ahead end-of-month CIC level RMSE.

    For each origin = last trading day of month M (level L_M known):
      1. Fit all 5 models on data up to origin (expanding window).
      2. Forecast all business days of M+1 via generate_future_exog('Daily_Baseline').
      3. EOM level forecast = L_M + Σ ΔCIC_hat over M+1.
      4. Error = actual EOM(M+1) level − forecast.

    All five ARIMAX/StateSpace variants share the same calendar dummy matrix so
    the same X_future is reused for all models.
    """
    arimax_models = ['Daily_Baseline']
    ss_models     = ['Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']
    all_keys      = arimax_models + ss_models
    store         = {k: {'dates': [], 'actual': [], 'forecast': []} for k in all_keys}
    comp_store    = {k: {'dates': [], 'comp1': [], 'comp2': [], 'total_fc': [], 'actual': []}
                     for k in all_keys}

    origins = pd.date_range(f'{start_year - 1}-12-31', f'{end_year - 1}-12-31', freq='ME')

    for origin in origins:
        avail = df.index[df.index <= origin]
        if len(avail) < 500:
            continue
        train_end = avail[-1]
        df_train  = df.loc[:train_end]

        nm_start = origin + pd.offsets.MonthBegin(1)
        nm_end   = nm_start + pd.offsets.MonthEnd(0)
        df_next  = df.loc[nm_start:nm_end]
        if len(df_next) < 5:
            continue

        lev_next = df_next['Currency'].dropna()
        if len(lev_next) == 0:
            continue
        actual_eom = float(lev_next.iloc[-1])
        lev_hist   = df_train['Currency'].dropna()
        if len(lev_hist) == 0:
            continue
        last_level = float(lev_hist.iloc[-1])

        fc_start = df_next.index[0].strftime('%Y-%m-%d')
        fc_end   = df_next.index[-1].strftime('%Y-%m-%d')
        try:
            X_fut_df = generate_future_exog('Daily_Baseline', fc_start, fc_end, hol,
                                            actual_dates=df_next.index)
        except Exception:
            continue
        if len(X_fut_df) == 0:
            continue

        y_chg = df_train['Change'].values

        # ARIMAX models
        for mname in arimax_models:
            X_tr, _ = get_X(df_train, mname)
            try:
                mdl    = TwoStepARIMAX().fit(y_chg, X_tr)
                fc     = mdl.forecast(X_fut_df.values)
                fc_eom = last_level + float(fc.sum())
                store[mname]['dates'].append(nm_end)
                store[mname]['actual'].append(actual_eom)
                store[mname]['forecast'].append(fc_eom)
                # comp1 = dummies contribution (no intercept); comp2 = intercept + ARIMA tail
                Xf_arr  = X_fut_df.values
                n_fc    = len(Xf_arr)
                c1_sum  = float((Xf_arr @ mdl.ols.coef_).sum())
                arima_fc_sum = float(np.asarray(mdl.arima.forecast(steps=n_fc)).sum())
                c2_sum  = float(mdl.ols.intercept_) * n_fc + arima_fc_sum
                comp_store[mname]['dates'].append(nm_end)
                comp_store[mname]['comp1'].append(last_level + c1_sum)
                comp_store[mname]['comp2'].append(last_level + c2_sum)
                comp_store[mname]['total_fc'].append(fc_eom)
                comp_store[mname]['actual'].append(actual_eom)
            except Exception as exc:
                print(f'    ⚠ {mname} EOM {origin.date()}: {exc}')

        # Daily_AdaptiveDrift — uses Daily_Baseline regressors
        X_tr_ss, _ = get_X(df_train, 'Daily_Baseline')
        try:
            mdl    = AdaptiveDriftModel('Daily_AdaptiveDrift').fit(y_chg, X_tr_ss)
            fc     = mdl.forecast(X_fut_df.values)
            fc_eom = last_level + float(fc.sum())
            store['Daily_AdaptiveDrift']['dates'].append(nm_end)
            store['Daily_AdaptiveDrift']['actual'].append(actual_eom)
            store['Daily_AdaptiveDrift']['forecast'].append(fc_eom)
            # comp1 = calendar dummies; comp2 = intercept + UC adaptive drift
            Xf_arr   = X_fut_df.values
            n_fc     = len(Xf_arr)
            c1_sum   = float((Xf_arr @ mdl.ols.coef_).sum())
            uc_fc_s  = float(np.asarray(mdl.uc_res.get_forecast(steps=n_fc).predicted_mean).sum())
            c2_sum   = float(mdl.ols.intercept_) * n_fc + uc_fc_s
            comp_store['Daily_AdaptiveDrift']['dates'].append(nm_end)
            comp_store['Daily_AdaptiveDrift']['comp1'].append(last_level + c1_sum)
            comp_store['Daily_AdaptiveDrift']['comp2'].append(last_level + c2_sum)
            comp_store['Daily_AdaptiveDrift']['total_fc'].append(fc_eom)
            comp_store['Daily_AdaptiveDrift']['actual'].append(actual_eom)
        except Exception as exc:
            print(f'    ⚠ Daily_AdaptiveDrift EOM {origin.date()}: {exc}')

        # Daily_AdaptiveSeasonal — uses Daily_Baseline regressors with trailing-window OLS
        X_tr_m3, _ = get_X(df_train, 'Daily_Baseline')
        try:
            mdl    = AdaptiveSeasonalModel().fit(y_chg, X_tr_m3, dates=df_train.index)
            fc     = mdl.forecast(X_fut_df.values)
            fc_eom = last_level + float(fc.sum())
            store['Daily_AdaptiveSeasonal']['dates'].append(nm_end)
            store['Daily_AdaptiveSeasonal']['actual'].append(actual_eom)
            store['Daily_AdaptiveSeasonal']['forecast'].append(fc_eom)
            # comp1 = calendar dummies (trailing-window betas); comp2 = intercept + UC drift
            Xf_arr   = X_fut_df.values
            n_fc     = len(Xf_arr)
            c1_sum   = float((Xf_arr @ mdl.ols.coef_).sum())
            uc_fc_s  = float(np.asarray(mdl.uc_res.get_forecast(steps=n_fc).predicted_mean).sum())
            c2_sum   = float(mdl.ols.intercept_) * n_fc + uc_fc_s
            comp_store['Daily_AdaptiveSeasonal']['dates'].append(nm_end)
            comp_store['Daily_AdaptiveSeasonal']['comp1'].append(last_level + c1_sum)
            comp_store['Daily_AdaptiveSeasonal']['comp2'].append(last_level + c2_sum)
            comp_store['Daily_AdaptiveSeasonal']['total_fc'].append(fc_eom)
            comp_store['Daily_AdaptiveSeasonal']['actual'].append(actual_eom)
        except Exception as exc:
            print(f'    ⚠ Daily_AdaptiveSeasonal EOM {origin.date()}: {exc}')

    # Finalise store
    for k in all_keys:
        r = store[k]
        r['dates']    = pd.DatetimeIndex(r['dates'])
        r['actual']   = np.array(r['actual'],   dtype=float)
        r['forecast'] = np.array(r['forecast'], dtype=float)
        r['errors']   = r['actual'] - r['forecast']
        r['RMSE']     = np.sqrt(np.mean(r['errors'] ** 2)) if len(r['errors']) > 0 else np.nan

    # Finalise comp_store
    for k in all_keys:
        c = comp_store[k]
        c['dates']    = pd.DatetimeIndex(c['dates'])
        c['comp1']    = np.array(c['comp1'],    dtype=float)
        c['comp2']    = np.array(c['comp2'],    dtype=float)
        c['total_fc'] = np.array(c['total_fc'], dtype=float)
        c['actual']   = np.array(c['actual'],   dtype=float)
    store['_comp'] = comp_store

    # Print summary
    print(f'\n  EOM Level Backtest  ({start_year}–{end_year}):')
    print(f'  {"Model":<20}  {"Overall RMSE":>14}  {"n":>5}')
    print('  ' + '-' * 42)
    for k in all_keys:
        r = store[k]
        print(f'  {k:<20}  {r["RMSE"]:>14.3f}  {len(r["errors"]):>5}')

    report_years = list(range(start_year, end_year + 1))
    print(f'\n  Per-year EOM RMSE:')
    hdr = f'  {"Model":<20}' + ''.join(f'{y:>8}' for y in report_years)
    print(hdr)
    print('  ' + '-' * (20 + 8 * len(report_years)))
    for k in all_keys:
        r   = store[k]
        row = f'  {k:<20}'
        for yr in report_years:
            mask = r['dates'].year == yr
            row += f'{np.sqrt(np.mean(r["errors"][mask]**2)):>8.2f}' \
                   if mask.sum() > 0 else f'{"—":>8}'
        print(row)

    return store


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7 — GARCH
# ─────────────────────────────────────────────────────────────────────────────

def fit_garch(residuals, label='GARCH(1,1)'):
    am  = arch_model(residuals, vol='GARCH', p=1, q=1, dist='normal', rescale=False)
    res = am.fit(disp='off', show_warning=False)
    p   = res.params
    print(f'  [{label}]  AIC={res.aic:.1f}  BIC={res.bic:.1f}  '
          f'ω={p["omega"]:.5f}  α={p["alpha[1]"]:.4f}  β={p["beta[1]"]:.4f}  '
          f'persistence={p["alpha[1]"]+p["beta[1]"]:.4f}')
    return res


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8 — VISUALIZATION
# ─────────────────────────────────────────────────────────────────────────────

def _save(fig, directory, filename):
    path = os.path.join(directory, filename)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'  Saved → ./{filename}')


def plot_data_overview(df, save_dir='.'):
    df_lev = df[df['Currency'].notna()]
    fig, axes = plt.subplots(2, 1, figsize=(15, 8))
    ax = axes[0]
    ax.plot(df_lev.index, df_lev['Currency'], color='#1f77b4', lw=0.9)
    ax.axvspan(pd.Timestamp('2020-03-01'), pd.Timestamp('2020-12-31'),
               alpha=0.12, color='red', label='COVID period')
    ax.axvline(pd.Timestamp('2020-03-24'), color='red', lw=1.2, ls='--', alpha=0.7,
               label='COVID 4-day dummy (D_Covid_1st)')
    ax.set_ylabel('CIC Level (THB billion)', fontsize=11)
    ax.set_title('Currency in Circulation — Daily Level (1997–2026)', fontsize=13, fontweight='bold')
    ax.legend(fontsize=9)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax = axes[1]
    ax.plot(df.index, df['Change'], color='#2ca02c', lw=0.55, alpha=0.8)
    ax.axhline(0, color='black', lw=0.5, ls='--')
    ax.axvspan(pd.Timestamp('2020-03-01'), pd.Timestamp('2020-12-31'),
               alpha=0.12, color='red', label='COVID period')
    ax.set_ylabel('Daily Change (THB billion)', fontsize=11)
    ax.set_title('Daily Change in CIC — Model Dependent Variable', fontsize=13, fontweight='bold')
    ax.legend(fontsize=9)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    fig.tight_layout(pad=2)
    _save(fig, save_dir, 'fig1_data_overview.png')


def plot_residual_diagnostics(residuals_dict, train_label, save_dir='.'):
    models = list(residuals_dict.keys())
    n      = len(models)
    fig, axes = plt.subplots(n, 2, figsize=(14, 4.2 * n))
    if n == 1:
        axes = axes[np.newaxis, :]
    for i, mname in enumerate(models):
        res  = np.asarray(residuals_dict[mname], float)
        res  = res[~np.isnan(res)]
        conf = 1.96 / np.sqrt(len(res))
        acf_vals = acf(res, nlags=min(40, len(res)//5), fft=True)
        col  = COLORS.get(mname, 'grey')
        ax   = axes[i, 0]
        ax.bar(range(len(acf_vals)), acf_vals, color=col, alpha=0.7)
        ax.axhline(conf,  color='red', ls='--', lw=0.8)
        ax.axhline(-conf, color='red', ls='--', lw=0.8)
        ax.axhline(0, color='black', lw=0.5)
        ax.set_title(f'{BASE_LABELS.get(mname, mname)}\nResidual ACF', fontsize=10)
        ax.set_xlabel('Lag')
        ax = axes[i, 1]
        (osm, osr), (slope, intercept, _) = stats.probplot(res, dist='norm')
        ax.scatter(osm, osr, s=6, alpha=0.5, color=col)
        ax.plot(osm, slope * np.array(osm) + intercept, 'r-', lw=1.5)
        ax.set_title(f'{BASE_LABELS.get(mname, mname)}\nNormal Q-Q', fontsize=10)
        ax.set_xlabel('Theoretical quantiles')
        ax.set_ylabel('Sample quantiles')
    fig.tight_layout(pad=2)
    _save(fig, save_dir, 'fig2_residual_diagnostics.png')


def plot_garch_volatility(train_index, residuals, garch_res, save_dir='.'):
    fig, axes = plt.subplots(2, 1, figsize=(15, 8), sharex=True)
    n   = len(residuals)
    idx = train_index[:n]
    ax  = axes[0]
    ax.plot(idx, residuals, color='steelblue', lw=0.6, alpha=0.85)
    ax.axhline(0, color='black', lw=0.5)
    ax.axvspan(pd.Timestamp('2020-03-01'), pd.Timestamp('2020-12-31'),
               alpha=0.15, color='red', label='COVID 2020')
    ax.set_title('Daily_Baseline — Training Residuals', fontsize=12, fontweight='bold')
    ax.set_ylabel('Residual (THB bn)')
    ax.legend(fontsize=9)
    ax = axes[1]
    cv  = garch_res.conditional_volatility[:len(idx)]
    ax.plot(idx[:len(cv)], cv, color='#d62728', lw=1.2)
    ax.fill_between(idx[:len(cv)], cv, alpha=0.2, color='#d62728',
                    label='Conditional σ (GARCH)')
    ax.axvspan(pd.Timestamp('2020-03-01'), pd.Timestamp('2020-12-31'),
               alpha=0.15, color='red')
    ax.set_title('GARCH(1,1) — Conditional Volatility', fontsize=12, fontweight='bold')
    ax.set_ylabel('Conditional Std Dev (THB bn)')
    ax.legend(fontsize=9)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    fig.tight_layout()
    _save(fig, save_dir, 'fig3_garch_volatility.png')


def plot_seasonal_pattern(df, fitted_models_dict, hol, save_dir='.'):
    """
    Seasonal CIC pattern: monthly end-of-month CIC level by year.
    Y-axis  : CIC level (THB billion)
    X-axis  : Month (Jan–Dec)
    Lines   : each year (last 10 years highlighted, older years faded)
    Fan     : 3-model forecast fan — Daily_Baseline / Daily_AdaptiveDrift /
              Daily_AdaptiveSeasonal — with shaded range across models
    """
    df_lev = df[df['Currency'].notna()].copy()
    eom    = df_lev['Currency'].resample('ME').last().dropna()

    pivot = pd.DataFrame({'month': eom.index.month,
                          'year':  eom.index.year,
                          'cic':   eom.values})
    pivot = pivot.pivot(index='month', columns='year', values='cic')

    recent_years = sorted([y for y in pivot.columns if y >= pivot.columns.max() - 9])
    n_yr = len(recent_years)

    fig, ax = plt.subplots(figsize=(14, 7))
    cmap_colors = cm.tab10(np.linspace(0, 0.9, n_yr))

    for i, yr in enumerate(recent_years):
        col_data = pivot.get(yr)
        if col_data is None:
            continue
        valid = col_data.dropna()
        style = {'lw': 2.0 if yr == recent_years[-1] else 1.2,
                 'alpha': 1.0 if yr >= recent_years[-2] else 0.65}
        ax.plot(valid.index, valid.values,
                color=cmap_colors[i], marker='o', ms=4,
                label=str(yr), **style)

    # Fan chart: 3 model forecasts shown as lines + shaded range
    last_date = df_lev.index.max()
    last_cic  = df_lev['Currency'].iloc[-1]
    fc_start  = (last_date + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
    fc_end    = (last_date + pd.Timedelta(days=45)).strftime('%Y-%m-%d')

    fan_models = {k: v for k, v in fitted_models_dict.items()
                  if k in ('Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal')}
    fan_fc = {}  # mname -> {(yr, mo): eom_level}

    for mname, mdl in fan_models.items():
        try:
            X_fut = generate_future_exog('Daily_Baseline' if mname in ('Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal') else mname,
                                         fc_start, fc_end, hol)
            if len(X_fut) == 0:
                continue
            fc_change = mdl.forecast(X_fut.values)
            cic_fc = last_cic
            eom_fc = {}
            for dt, chg in zip(X_fut.index, fc_change):
                cic_fc += chg
                eom_fc[(dt.year, dt.month)] = cic_fc
            fan_fc[mname] = eom_fc
        except Exception as e:
            print(f'  (Fan forecast skipped for {mname}: {e})')

    if fan_fc:
        # Shaded range between min/max forecast across models
        all_keys = set()
        for eom_fc in fan_fc.values():
            all_keys.update(eom_fc.keys())
        for yr_mo in sorted(all_keys):
            mo = yr_mo[1]
            vals = [fan_fc[m][yr_mo] for m in fan_models if yr_mo in fan_fc.get(m, {})]
            if len(vals) >= 2:
                ax.fill_between([mo - 0.3, mo + 0.3],
                                [min(vals), min(vals)], [max(vals), max(vals)],
                                color='#cccccc', alpha=0.5, zorder=7)

        fan_label_map = {k: f'{k} fc' for k in
                         ('Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal')}
        for mname in ('Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal'):
            eom_fc = fan_fc.get(mname, {})
            if not eom_fc:
                continue
            xs = [k[1] for k in sorted(eom_fc)]
            ys = [eom_fc[k] for k in sorted(eom_fc)]
            ax.plot(xs, ys, marker='o', ms=7, lw=0, zorder=10,
                    color=COLORS.get(mname, 'grey'),
                    markeredgecolor='black', markeredgewidth=0.8,
                    label=fan_label_map[mname])

    month_names = ['Jan','Feb','Mar','Apr','May','Jun',
                   'Jul','Aug','Sep','Oct','Nov','Dec']
    ax.set_xticks(range(1, 13))
    ax.set_xticklabels(month_names, fontsize=11)
    ax.set_ylabel('CIC Level (THB billion)', fontsize=11)
    ax.set_title('Seasonal CIC Pattern — End-of-Month Level by Year\n'
                 '(dots = next-month forecast; shaded band = range across Daily_Baseline / Daily_AdaptiveDrift / Daily_AdaptiveSeasonal)',
                 fontsize=13, fontweight='bold')
    ax.legend(fontsize=8.5, ncol=1, loc='upper left',
              bbox_to_anchor=(1.01, 1), borderaxespad=0,
              title='Year / Model', title_fontsize=9)
    ax.grid(alpha=0.3)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))
    fig.tight_layout()
    _save(fig, save_dir, 'fig4_seasonal_pattern.png')


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8d — CIC OUTPUT EXCEL (single consolidated user-facing workbook)
# ─────────────────────────────────────────────────────────────────────────────
# CIC_output.xlsx is the one Excel deliverable: the daily pipeline (below) and
# the --eom harness (export_results(), Part 2) both write into it. Each writer
# only ever replaces the sheets it owns — daily pipeline: Daily / Monthly EOM /
# Summary; --eom harness: EOM_Selection / EOM_Holdout / EOM_Detail /
# Daily_Guardrail — so running one after the other accumulates, it never wipes
# the other's sheets. Whichever runs first creates the file.

def _excel_writer(path):
    """pd.ExcelWriter that appends/replaces sheets in an existing workbook
    instead of truncating it, so two independent scripts can share one file."""
    if os.path.exists(path):
        return pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    return pd.ExcelWriter(path, engine='openpyxl', mode='w')


def export_cic_output_excel(df, configs_results, hol, save_dir='.'):
    """
    CIC_output.xlsx — 3 tabs:

    Daily       : date | CIC actual | one column per model (CIC level, 1-step-ahead)
                  OOS 2020-present + 2 months forward; yellow = forecast rows

    Monthly EOM : date | CIC actual | CIC actual change | one EOM column per model
                  Yellow = forecast rows

    Summary     : rows = one per model + Avg Post-COVID Seasonal (SUMPRODUCT formula)
                  cols = next 2 forecast months
                  values = monthly CIC change
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    YELLOW  = PatternFill(start_color='FFFFC0', end_color='FFFFC0', fill_type='solid')
    BLUE_HD = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    GREY_HD = PatternFill(start_color='595959', end_color='595959', fill_type='solid')
    HDR_FONT = Font(color='FFFFFF', bold=True)
    BOLD     = Font(bold=True)

    path = os.path.join(save_dir, 'CIC_output.xlsx')

    # ── Source data ──
    main = configs_results['cfg_main']
    df_train   = main['df_train']
    df_eval    = main['df_eval']
    forecasts  = main['forecasts']          # mname -> np.array of daily change forecasts

    # Display order: baseline first, then the adaptive variants
    MODEL_ORDER = ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']

    last_actual     = df['Currency'].dropna().index.max()
    last_actual_eom = df['Currency'].dropna().resample('ME').last().dropna().index.max()
    last_train_lv   = df_train['Currency'].dropna().iloc[-1]

    # ── 1-step-ahead daily CIC level for each model (OOS period) ──
    # level_t = actual_{t-1} + model_forecast_change_t
    actual_prev = df['Currency'].shift(1)

    oos_level = {}
    for mname in MODEL_ORDER:
        fc_arr = forecasts.get(mname)
        if fc_arr is None:
            continue
        fc_s = pd.Series(fc_arr, index=df_eval.index)
        oos_level[mname] = actual_prev[df_eval.index] + fc_s

    # ── 2-month future daily forecasts (cumulative from last actual) ──
    fc_start = (last_actual + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
    fc_end   = (last_actual_eom + pd.offsets.MonthEnd(2)).strftime('%Y-%m-%d')  # 2 full future months

    future_fc    = {}   # mname -> pd.Series(change, future dates)
    future_level = {}   # mname -> pd.Series(level, future dates)
    for mname in MODEL_ORDER:
        mdl = main['fitted_models'].get(mname)
        if mdl is None:
            continue
        try:
            key = 'Daily_Baseline' if mname in ('Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal') else mname
            X_fut = generate_future_exog(key, fc_start, fc_end, hol)
            if not len(X_fut):
                continue
            fc_chg = mdl.forecast(X_fut.values)
            future_fc[mname] = pd.Series(fc_chg, index=X_fut.index)
            lvl = df['Currency'].dropna().iloc[-1]
            lvls = []
            for chg in fc_chg:
                lvl += chg
                lvls.append(lvl)
            future_level[mname] = pd.Series(lvls, index=X_fut.index)
        except Exception as e:
            print(f'  (future fc skipped {mname}: {e})')

    # ── Build Daily DataFrame ──
    # Rows: OOS 2020-present + future 2 months
    actual_oos = df.loc[df_eval.index, 'Currency']
    daily_rows = []
    # OOS actual rows
    for dt in df_eval.index:
        row = {'Date': dt, 'CIC Actual (bn.)': actual_oos.get(dt, np.nan)}
        for mname in MODEL_ORDER:
            lbl = f'CIC {BASE_LABELS.get(mname, mname)}'
            row[lbl] = oos_level.get(mname, pd.Series(dtype=float)).get(dt, np.nan)
        daily_rows.append(row)
    # Future forecast rows
    if future_level:
        ref_idx = next(iter(future_level.values())).index
        for dt in ref_idx:
            row = {'Date': dt, 'CIC Actual (bn.)': np.nan}
            for mname in MODEL_ORDER:
                lbl = f'CIC {BASE_LABELS.get(mname, mname)}'
                row[lbl] = future_level.get(mname, pd.Series(dtype=float)).get(dt, np.nan)
            daily_rows.append(row)
    daily_df = pd.DataFrame(daily_rows)
    n_oos_rows = len(df_eval)  # forecast rows start after these

    # ── Build Monthly EOM DataFrame ──
    eom_actual = df['Currency'].dropna().resample('ME').last()
    eom_actual_chg = eom_actual.diff()

    # OOS EOM levels from 1-step-ahead forecasts
    eom_oos = {}
    for mname, lev_s in oos_level.items():
        eom_oos[mname] = lev_s.resample('ME').last()

    # Future EOM levels
    eom_future = {}
    for mname, lev_s in future_level.items():
        eom_future[mname] = lev_s.resample('ME').last()

    # Collect all months (actual + future)
    all_months = sorted(set(eom_actual.index) | set().union(*[s.index for s in eom_future.values()]) if eom_future else set(eom_actual.index))
    eom_rows = []
    for mo in all_months:
        is_fc = mo > last_actual
        row = {
            'Date': mo,
            'CIC Actual (bn.)': eom_actual.get(mo, np.nan),
            'CIC Actual Change (bn.)': eom_actual_chg.get(mo, np.nan),
        }
        for mname in MODEL_ORDER:
            lbl = f'CIC {BASE_LABELS.get(mname, mname)}'
            if is_fc:
                val = eom_future.get(mname, pd.Series(dtype=float)).get(mo, np.nan)
            else:
                val = eom_oos.get(mname, pd.Series(dtype=float)).get(mo, np.nan)
            row[lbl] = val
        eom_rows.append(row)
    eom_df = pd.DataFrame(eom_rows)
    n_actual_eom = int((eom_df['Date'] <= last_actual).sum())

    # ── Build Summary DataFrame ──
    # Columns = next 2 forecast months, Rows = 2 models + avg seasonal
    summary_models   = ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']
    summary_row_lbls = [BASE_LABELS.get(m, m) for m in summary_models] + ['Avg Post-COVID Seasonal']

    # Identify the 2 forecast month-end dates (strictly after last actual EOM)
    last_actual_eom = eom_actual.dropna().index.max()
    fc_months = sorted(
        m for m in (eom_future[next(iter(eom_future))].index if eom_future else [])
        if m > last_actual_eom
    )[:2]

    # For each model: monthly change = forecast EOM level - previous EOM level
    def eom_change_for_month(mname, mo):
        prev_mos = [m for m in all_months if m < mo]
        if not prev_mos:
            return np.nan
        prev_mo = prev_mos[-1]
        cur_lev  = eom_future.get(mname, pd.Series(dtype=float)).get(mo, np.nan)
        prev_lev_fc  = eom_future.get(mname, pd.Series(dtype=float)).get(prev_mo, np.nan)
        prev_lev_act = eom_actual.get(prev_mo, np.nan)
        prev_lev = prev_lev_fc if not np.isnan(prev_lev_fc) else prev_lev_act
        return cur_lev - prev_lev if not np.isnan(cur_lev) and not np.isnan(prev_lev) else np.nan

    summary_data = {}
    for mo in fc_months:
        col_lbl = mo.strftime('%b %Y')
        col_vals = []
        for mname in summary_models:
            col_vals.append(eom_change_for_month(mname, mo))
        col_vals.append(None)  # placeholder for formula row
        summary_data[col_lbl] = col_vals

    summary_df = pd.DataFrame(summary_data, index=summary_row_lbls)
    summary_df.index.name = 'Model'

    # ── Write to Excel (append-safe: preserves any EOM_* sheets from --eom) ──
    with _excel_writer(path) as writer:
        daily_df.to_excel(writer, sheet_name='Daily', index=False)
        eom_df.to_excel(writer, sheet_name='Monthly EOM', index=False)
        summary_df.reset_index().to_excel(writer, sheet_name='Summary', index=False)

    # ── Style with openpyxl ──
    wb = load_workbook(path)

    def style_sheet(ws, forecast_start_row, n_cols, freeze_col=1):
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill      = BLUE_HD
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            is_fc = row[0].row >= forecast_start_row
            for cell in row:
                if is_fc:
                    cell.fill = YELLOW
                cell.border = border
                if cell.column == 1:
                    cell.number_format = 'YYYY-MM-DD'
                elif cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 28)
        ws.freeze_panes = ws.cell(row=2, column=freeze_col + 1)

    # Daily
    style_sheet(wb['Daily'], n_oos_rows + 2, daily_df.shape[1])

    # Monthly EOM
    style_sheet(wb['Monthly EOM'], n_actual_eom + 2, eom_df.shape[1])

    # Summary — custom styling + SUMPRODUCT formulas for last row
    ws_s = wb['Summary']
    n_s_cols = summary_df.shape[1] + 1  # +1 for index col

    # Style header
    for col in range(1, n_s_cols + 1):
        cell = ws_s.cell(row=1, column=col)
        cell.fill = GREY_HD
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center')

    # Style data rows
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws_s.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.font = BOLD
            elif cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # Write SUMPRODUCT formula for "Avg Post-COVID Seasonal" row
    # Last data row in summary = row (len(summary_row_lbls) + 1) = row 5
    avg_row = len(summary_row_lbls) + 1  # row index in worksheet (1-based, +1 for header)
    # Monthly EOM tab: col A = Date, col C = CIC Actual Change
    # fc_months[i] is in col (i+2) of summary (col B, C, ...)
    for col_idx, mo in enumerate(fc_months, start=2):
        mo_num = mo.month
        # =SUMPRODUCT((MONTH('Monthly EOM'!$A$2:$A$500)=mo_num)*
        #             ('Monthly EOM'!$A$2:$A$500>=DATE(2021,1,1))*
        #             ('Monthly EOM'!$A$2:$A$500<=DATE(2025,12,31))*
        #             ISNUMBER('Monthly EOM'!$C$2:$C$500)*
        #             ('Monthly EOM'!$C$2:$C$500))
        # / SUMPRODUCT(...)
        rng_a = "'Monthly EOM'!$A$2:$A$500"
        rng_c = "'Monthly EOM'!$C$2:$C$500"
        cond  = (f"(MONTH({rng_a})={mo_num})*"
                 f"({rng_a}>=DATE(2021,1,1))*"
                 f"({rng_a}<=DATE(2025,12,31))*"
                 f"ISNUMBER({rng_c})")
        formula = (f"=SUMPRODUCT({cond}*({rng_c}))"
                   f"/SUMPRODUCT({cond}*1)")
        cell = ws_s.cell(row=avg_row, column=col_idx)
        cell.value = formula
        cell.number_format = '#,##0.00'
        cell.fill   = YELLOW
        cell.border = border

    # Auto-width summary
    for col in ws_s.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_s.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    wb.save(path)
    print(f'  Saved → ./CIC_output.xlsx')


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9 — EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(df, configs_results, rolling_metrics, h_rmse, garch_res,
                 eom_results=None, save_dir='.'):
    """
    Export results to cic_forecast_output.xlsx

    Sheets
    ------
    Eval_Benchmark  – Dec 2021–May 2022 eval rows only (all models, easy to read)
    Eval_Extended   – Jan 2024–Dec 2025 eval rows only (extended config)
    InSample_Fitted – in-sample fitted values for each model (benchmark config)
    Full_Series     – complete daily CIC level + change (no forecasts, for reference)
    Benchmark_Metrics – RMSE/MAE/ResidSD/Bias for all model-config combos
    Rolling_RMSE    – expanding-window backtest
    Horizon_RMSE    – multi-horizon RMSE
    GARCH_Params    – GARCH(1,1) estimates
    """
    path = os.path.join(save_dir, 'cic_forecast_output.xlsx')

    with pd.ExcelWriter(path, engine='openpyxl') as writer:

        # ── Eval sheets: one row per eval observation, all models side by side ──
        for cfg_key, cfg_data in configs_results.items():
            df_eval     = cfg_data['df_eval']
            forecasts   = cfg_data['forecasts']
            fitted_mdls = cfg_data['fitted_models']
            train_lbl   = cfg_data['train_label']

            out = pd.DataFrame({'Date': df_eval.index,
                                'CIC_Level': df_eval['Currency'].values,
                                'Change_Actual': df_eval['Change'].values})
            for mname, pred in forecasts.items():
                lbl = model_label(mname, train_lbl)
                out[f'{lbl}_Change_Forecast'] = pred
                # Reconstruct CIC level forecast
                last_cic  = df[df.index < df_eval.index[0]]['Currency'].iloc[-1]
                cic_levels = np.empty(len(pred))
                prev = last_cic
                for j, chg in enumerate(pred):
                    prev += chg
                    cic_levels[j] = prev
                out[f'{lbl}_CIC_Forecast'] = cic_levels

            sheet_name = 'Eval_Benchmark' if cfg_key == 'cfg_benchmark' else 'Eval_Main'
            out.to_excel(writer, sheet_name=sheet_name, index=False)

        # ── In-sample fitted values (main config) ──
        for ck in ['cfg_main', 'cfg_benchmark']:
            if ck not in configs_results:
                continue
            cdata   = configs_results[ck]
            df_tr   = cdata['df_train']
            lbl     = cdata['train_label']
            isfit   = pd.DataFrame({'Date': df_tr.index,
                                    'CIC_Level': df_tr['Currency'].values,
                                    'Change_Actual': df_tr['Change'].values})
            for mname, mdl in cdata['fitted_models'].items():
                fitted_chg = np.asarray(mdl.fitted)
                n_fit = min(len(fitted_chg), len(df_tr))
                col   = np.full(len(df_tr), np.nan)
                col[:n_fit] = fitted_chg[:n_fit]
                col_label = model_label(mname, lbl)
                isfit[f'{col_label}_Fitted'] = col
            sname = 'InSample_Main' if ck == 'cfg_main' else 'InSample_Benchmark'
            isfit.to_excel(writer, sheet_name=sname, index=False)
            break  # only write one

        # ── Full series ──
        full = df[['Currency', 'Change']].copy()
        full.columns = ['CIC_Level', 'Change']
        full.index.name = 'Date'
        full.reset_index().to_excel(writer, sheet_name='Full_Series', index=False)

        # ── Benchmark metrics (all model-config combos) ──
        rows = []
        for cfg_key, cfg_data in configs_results.items():
            lbl = cfg_data['train_label']
            ev  = cfg_data['eval_label']
            for mname, m in cfg_data['bench_metrics'].items():
                rows.append({
                    'Model': model_label(mname, lbl),
                    'Eval window': ev,
                    'RMSE (THB bn)': round(m['RMSE'],    4),
                    'MAE (THB bn)':  round(m['MAE'],     4),
                    'ResidSD':       round(m['ResidSD'], 4),
                    'Bias':          round(m['Bias'],    4),
                    'n obs':         m['n'],
                })
        rows.append({'Model': '[BOT 2022 paper (2017-2021)]',
                     'Eval window': 'Dec 2021–May 2022',
                     'RMSE (THB bn)': 4.960, 'MAE (THB bn)': None,
                     'ResidSD': 4.140, 'Bias': None, 'n obs': 119})
        rows.append({'Model': '[Pre-2022 model]',
                     'Eval window': 'Dec 2021–May 2022',
                     'RMSE (THB bn)': 7.310, 'MAE (THB bn)': None,
                     'ResidSD': 4.750, 'Bias': None, 'n obs': 119})
        pd.DataFrame(rows).to_excel(writer, sheet_name='Benchmark_Metrics', index=False)

        # ── Rolling RMSE ──
        roll_rows = []
        all_win = []
        for mname, wdict in rolling_metrics.items():
            all_win = list(wdict.keys())
            break
        for mname, wdict in rolling_metrics.items():
            row = {'Model': BASE_LABELS[mname]}
            for wl in all_win:
                val = wdict.get(wl, {}).get('RMSE', np.nan)
                row[wl] = round(float(val), 4) if not np.isnan(val) else None
            roll_rows.append(row)
        pd.DataFrame(roll_rows).to_excel(writer, sheet_name='Rolling_RMSE', index=False)

        # ── Horizon RMSE ──
        h_rows = []
        for mname, hdict in h_rmse.items():
            row = {'Model': BASE_LABELS[mname]}
            for h in [1, 5, 10, 22]:
                val = hdict.get(h, np.nan)
                row[f'h={h}d'] = round(float(val), 4) if not np.isnan(val) else None
            h_rows.append(row)
        pd.DataFrame(h_rows).to_excel(writer, sheet_name='Horizon_RMSE', index=False)

        # ── GARCH ──
        gp = garch_res.params
        garch_rows = [
            {'Parameter': 'omega',       'Value': round(float(gp['omega']),    6)},
            {'Parameter': 'alpha[1]',    'Value': round(float(gp['alpha[1]']), 6)},
            {'Parameter': 'beta[1]',     'Value': round(float(gp['beta[1]']),  6)},
            {'Parameter': 'persistence', 'Value': round(float(gp['alpha[1]'] + gp['beta[1]']), 6)},
            {'Parameter': 'AIC',         'Value': round(float(garch_res.aic), 2)},
            {'Parameter': 'BIC',         'Value': round(float(garch_res.bic), 2)},
        ]
        pd.DataFrame(garch_rows).to_excel(writer, sheet_name='GARCH_Params', index=False)

        # ── Level EOM Metrics (Model D backtest) ──
        if eom_results:
            eom_rows = []
            model_order = ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']
            for k in model_order:
                r = eom_results.get(k, {})
                if not len(r.get('dates', [])):
                    continue
                dates  = r['dates']
                actual = r['actual']
                fc     = r['forecast']
                errs   = r['errors']
                years  = sorted(set(dates.year))
                base_row = {'Model': k, 'Overall_RMSE': round(r['RMSE'], 3), 'n': len(errs)}
                for yr in years:
                    mask = dates.year == yr
                    if mask.sum() > 0:
                        base_row[f'RMSE_{yr}'] = round(
                            float(np.sqrt(np.mean(errs[mask] ** 2))), 3)
                eom_rows.append(base_row)
            pd.DataFrame(eom_rows).to_excel(writer, sheet_name='Level_EOM_Metrics', index=False)

            # Also write detail rows (date, actual, forecast, error per model)
            detail_frames = []
            for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']:
                r = eom_results.get(k, {})
                if not len(r.get('dates', [])):
                    continue
                tmp = pd.DataFrame({
                    'Model':    k,
                    'Date':     r['dates'],
                    'Actual_EOM_Level':   r['actual'],
                    'Forecast_EOM_Level': r['forecast'],
                    'Error':              r['errors'],
                })
                detail_frames.append(tmp)
            if detail_frames:
                pd.concat(detail_frames).sort_values(['Date', 'Model']).to_excel(
                    writer, sheet_name='Level_EOM_Detail', index=False)

            # ── Component Decomposition sheet ──
            comp_store = eom_results.get('_comp', {})
            if comp_store:
                comp_detail_frames = []
                for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']:
                    c = comp_store.get(k, {})
                    if not len(c.get('dates', [])):
                        continue
                    tmp = pd.DataFrame({
                        'Date':            c['dates'],
                        'Model':           k,
                        'Comp1_Calendar':  c['comp1'],
                        'Comp2_Drift':     c['comp2'],
                        'Total_Forecast':  c['total_fc'],
                        'Actual':          c['actual'],
                        'Error':           c['actual'] - c['total_fc'],
                    })
                    comp_detail_frames.append(tmp)
                if comp_detail_frames:
                    comp_df = pd.concat(comp_detail_frames).sort_values(['Date', 'Model'])
                    comp_df.to_excel(writer, sheet_name='Component_Decomp', index=False)

                    # Summary: RMSE(actual - comp_only_level) per model per year
                    # Comp1_RMSE = error using calendar dummies alone (no drift)
                    # Comp2_RMSE = error using drift alone (no calendar)
                    sum_rows = []
                    for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']:
                        c = comp_store.get(k, {})
                        if not len(c.get('dates', [])):
                            continue
                        act_c  = c['actual']
                        c1_arr = c['comp1']
                        c2_arr = c['comp2']
                        years = sorted(set(c['dates'].year))
                        row = {'Model': k}
                        for yr in years:
                            mask = c['dates'].year == yr
                            if mask.sum() > 0:
                                e1 = act_c[mask] - c1_arr[mask]
                                e2 = act_c[mask] - c2_arr[mask]
                                row[f'Comp1_CalOnly_RMSE_{yr}']   = round(float(np.sqrt(np.mean(e1**2))), 3)
                                row[f'Comp2_DriftOnly_RMSE_{yr}'] = round(float(np.sqrt(np.mean(e2**2))), 3)
                        sum_rows.append(row)
                    if sum_rows:
                        pd.DataFrame(sum_rows).to_excel(writer, sheet_name='Component_Summary', index=False)

    print(f'  Saved → ./cic_forecast_output.xlsx')


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10 — MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run_config(df, cfg, all_models):
    """Fit and evaluate all models for one train/eval configuration."""
    train_lbl  = cfg['train_label']
    train_end  = cfg['train_end']
    eval_start = cfg['eval_start']
    eval_end   = cfg['eval_end']

    df_train = df.loc[:train_end]
    df_eval  = df.loc[eval_start:eval_end]

    print(f'\n  Config ({train_lbl}):  '
          f'train {df_train.index[0].date()}→{df_train.index[-1].date()} '
          f'(n={len(df_train)}) | eval {eval_start}→{eval_end} (n={len(df_eval)})')

    fitted_models = {}
    for mname in all_models:
        X_tr, _ = get_X(df_train, mname)
        mdl = TwoStepARIMAX().fit(df_train['Change'].values, X_tr)
        fitted_models[mname] = mdl

    forecasts    = {}
    bench_metrics = {}
    actual_arr   = df_eval['Change'].values
    for mname, mdl in fitted_models.items():
        X_ev, _ = get_X(df_eval, mname)
        pred = mdl.forecast(X_ev)
        forecasts[mname]     = pred
        bench_metrics[mname] = compute_metrics(actual_arr, pred)

    return {
        'train_label':   train_lbl,
        'eval_label':    f'{eval_start[:7]}→{eval_end[:7]}',
        'df_train':      df_train,
        'df_eval':       df_eval,
        'fitted_models': fitted_models,
        'forecasts':     forecasts,
        'bench_metrics': bench_metrics,
    }


def run_full_pipeline():
    FILEPATH = 'input.xlsx'

    # Config 1 — Benchmark (BOT 2022 paper reference, kept for Excel comparison only)
    CFG_BENCHMARK = {
        'key':         'cfg_benchmark',
        'train_label': '1997-2021',
        'train_end':   '2021-11-30',
        'eval_start':  '2021-12-01',
        'eval_end':    '2022-05-31',
    }

    # Config 2 — Main: train pre-COVID (1997-2019), OOS = Jan 2020 – latest data
    # ~20% of 7000 obs is OOS; covers COVID, post-COVID, and 2024-25 trend shift
    CFG_MAIN = {
        'key':         'cfg_main',
        'train_label': '1997-2019',
        'train_end':   '2019-12-31',
        'eval_start':  '2020-01-01',
        'eval_end':    '2026-05-31',   # will be clipped to last available date
    }

    # Config 3 — Pre-COVID: train 1997-2017, OOS = 2018-2019 (pure out-of-sample)
    CFG_PRECOVID = {
        'key':         'cfg_precovid',
        'train_label': '1997-2017',
        'train_end':   '2017-12-31',
        'eval_start':  '2018-01-01',
        'eval_end':    '2019-12-31',
    }

    # Expanding-window rolling backtest: 1-year OOS each window, covers full history
    BACKTEST_WINDOWS = [
        ('2018-12-31', '2019-01-01', '2019-12-31'),
        ('2019-12-31', '2020-01-01', '2020-12-31'),
        ('2020-12-31', '2021-01-01', '2021-12-31'),
        ('2021-12-31', '2022-01-01', '2022-12-31'),
        ('2022-12-31', '2023-01-01', '2023-12-31'),
        ('2023-12-31', '2024-01-01', '2024-12-31'),
        ('2024-12-31', '2025-01-01', '2025-12-31'),
    ]
    # Horizon RMSE origins: spread across the OOS period
    HORIZON_ORIGINS = ['2020-01-01', '2022-06-01', '2024-01-01']
    ALL_MODELS      = list(REGS.keys())   # ['Daily_Baseline']
    CORE_MODELS     = ALL_MODELS

    sep = '=' * 65
    print(sep)
    print('  CIC FORECASTING — ORIGINAL vs. ADAPTIVE MODEL  (Bank of Thailand)')
    print(sep)

    # ── 1. Load ──
    print('\n[1] Loading data...')
    df  = load_data(FILEPATH)
    hol = load_holiday(FILEPATH)
    print(f'    Obs: {len(df)} | {df.index[0].date()} → {df.index[-1].date()}')
    chg = df['Change'].dropna()
    print(f'    Change: mean={chg.mean():.3f}  std={chg.std():.3f}  '
          f'min={chg.min():.3f}  max={chg.max():.3f} (THB bn)')
    adf_s, adf_p, *_ = adfuller(chg, autolag='AIC')
    print(f'    ADF: stat={adf_s:.3f}, p={adf_p:.4f}  → '
          f'{"stationary ✓" if adf_p<0.05 else "non-stationary ⚠"}')
    sk_tr = int(df.loc[:CFG_MAIN['train_end'], 'D_SK_PRE1'].sum())
    ny_tr = int(df.loc[:CFG_MAIN['train_end'], 'D_NY_PRE1'].sum())
    print(f'    D_SK_PRE1 (to {CFG_MAIN["train_end"][:4]}): {sk_tr} events  |  D_NY_PRE1: {ny_tr} events')

    # ── 2. Fig 1 ──
    print('\n[2] Figure 1 — CIC overview...')
    plot_data_overview(df)

    # ── 3. Fit all models on both configs (ARIMAX + Daily_AdaptiveDrift) ──
    print('\n[3] Fitting models on both training configurations...')
    print(f'  {"Model":<20} {"AIC":>10} {"BIC":>10} {"σ":>7} {"AR":>6} {"MA":>6}  [config]')
    print('  ' + '-' * 72)

    configs_results = {}
    for cfg in [CFG_BENCHMARK, CFG_MAIN, CFG_PRECOVID]:
        key      = cfg['key']
        lbl      = cfg['train_label']
        df_train = df.loc[:cfg['train_end']]
        # Clip eval end to last available date
        eval_end_actual = min(cfg['eval_end'], df.index[-1].strftime('%Y-%m-%d'))
        df_eval  = df.loc[cfg['eval_start']:eval_end_actual]

        # ARIMAX variants
        fitted_models = {}
        for mname in ALL_MODELS:
            X_tr, _ = get_X(df_train, mname)
            mdl = TwoStepARIMAX().fit(df_train['Change'].values, X_tr)
            fitted_models[mname] = mdl
            print(f'  [{model_label(mname, lbl):<28}]  '
                  f'AIC={mdl.aic:9.1f}  BIC={mdl.bic:9.1f}  '
                  f'σ={mdl.sigma:.3f}  AR={mdl.ar1:.3f}  MA={mdl.ma1:.3f}')

        # Daily_AdaptiveDrift state-space variant
        X_tr_ss, _ = get_X(df_train, 'Daily_Baseline')
        try:
            d1_mdl = AdaptiveDriftModel('Daily_AdaptiveDrift').fit(df_train['Change'].values, X_tr_ss)
            fitted_models['Daily_AdaptiveDrift'] = d1_mdl
            print(f'  [{"Daily_AdaptiveDrift ("+lbl+")":<28}]  AIC={d1_mdl.aic:9.1f}  BIC={d1_mdl.bic:9.1f}')
        except Exception as exc:
            print(f'  ⚠ Daily_AdaptiveDrift ({lbl}): {exc}')

        # Daily_AdaptiveSeasonal — smooth trend with trailing-window OLS
        X_tr_m3, _ = get_X(df_train, 'Daily_Baseline')
        try:
            m3_mdl = AdaptiveSeasonalModel().fit(
                df_train['Change'].values, X_tr_m3,
                dates=df_train.index
            )
            fitted_models['Daily_AdaptiveSeasonal'] = m3_mdl
            print(f'  [{"Daily_AdaptiveSeasonal ("+lbl+")":<28}]  AIC={m3_mdl.aic:9.1f}  BIC={m3_mdl.bic:9.1f}')
        except Exception as exc:
            print(f'  ⚠ Daily_AdaptiveSeasonal ({lbl}): {exc}')

        # Forecasts and metrics for eval period
        forecasts     = {}
        bench_metrics = {}
        actual_arr    = df_eval['Change'].values
        for mname, mdl in fitted_models.items():
            if mname in ('Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal'):
                X_ev, _ = get_X(df_eval, 'Daily_Baseline')
            else:
                X_ev, _ = get_X(df_eval, mname)
            pred = mdl.forecast(X_ev)
            forecasts[mname]     = pred
            bench_metrics[mname] = compute_metrics(actual_arr, pred)

        configs_results[key] = {
            'train_label':   lbl,
            'eval_label':    f'{cfg["eval_start"][:7]}→{eval_end_actual[:7]}',
            'df_train':      df_train,
            'df_eval':       df_eval,
            'fitted_models': fitted_models,
            'forecasts':     forecasts,
            'bench_metrics': bench_metrics,
        }

    ALL_MODELS_WITH_SS = ALL_MODELS + ['Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']  # ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']

    # ── 4. Benchmark metrics ──
    print('\n[4] Benchmark metrics:')
    for cfg_key, cfg_data in configs_results.items():
        lbl = cfg_data['train_label']
        ev  = cfg_data['eval_label']
        print(f'\n  Config ({lbl}) — eval {ev}:')
        print(f'  {"Model":<32} {"RMSE":>8} {"MAE":>8} {"ResidSD":>10}')
        print('  ' + '-' * 60)
        for mname, m in cfg_data['bench_metrics'].items():
            lbl2 = model_label(mname, lbl)
            print(f'  {lbl2:<32} {m["RMSE"]:>8.3f} {m["MAE"]:>8.3f} {m["ResidSD"]:>10.3f}')
        if cfg_key == 'cfg_benchmark':
            print(f'  {"[BOT 2022 paper (2017-2021)]":<32} {"4.960":>8} {"---":>8} {"4.140":>10}  (published)')
            print(f'  {"[Pre-2022 model]":<32} {"7.310":>8} {"---":>8} {"4.750":>10}  (published)')

    # ── 5. Residual diagnostics (main config, ARIMAX only) ──
    print('\n[5] Residual diagnostics (cfg_main training residuals)...')
    m_fitted  = configs_results['cfg_main']['fitted_models']
    m_lbl     = configs_results['cfg_main']['train_label']
    residuals = {k: v.resid for k, v in m_fitted.items() if k in ALL_MODELS}
    for mname, res in residuals.items():
        run_diagnostics(res, label=model_label(mname, m_lbl))

    # ── 6. ARCH + GARCH ──
    print('\n[6] ARCH-LM + GARCH(1,1) on Daily_Baseline residuals...')
    old_res = np.asarray(residuals['Daily_Baseline'], float)
    old_res = old_res[~np.isnan(old_res)]
    arch_stat, arch_pval, _, _ = het_arch(old_res, nlags=10)
    print(f'  ARCH-LM(10): stat={arch_stat:.3f}, p={arch_pval:.4f}  '
          f'→ {"⚠ ARCH effects present" if arch_pval<0.05 else "✓ no ARCH"}')
    garch_res = fit_garch(old_res)

    # ── 7. Rolling backtest (all 7 windows, Daily_Baseline) ──
    print('\n[7] Rolling backtest (expanding window, 7 periods, Daily_Baseline)...')
    rolling_metrics = rolling_backtest(df, CORE_MODELS, BACKTEST_WINDOWS)
    win_labels = [f'{es[:7]}→{ee[:7]}' for _, es, ee in BACKTEST_WINDOWS]
    print('\n  Rolling RMSE summary:')
    w_fmt = ''.join(f'{w:>16}' for w in win_labels)
    print(f'  {"Model":<24}{w_fmt}')
    print('  ' + '-' * (24 + 16 * len(win_labels)))
    for mname in CORE_MODELS:
        row = f'  {BASE_LABELS[mname]:<24}'
        for wl in win_labels:
            val = rolling_metrics.get(mname, {}).get(wl, {}).get('RMSE', np.nan)
            row += f'{val:>16.3f}' if not np.isnan(val) else f'{"—":>16}'
        print(row)

    # ── 8. Horizon RMSE ──
    print('\n[8] Horizon RMSE (1, 5, 10, 22-day ahead, 3 origins) — full OOS period...')
    h_rmse = horizon_rmse_monthly(df, CORE_MODELS, HORIZON_ORIGINS)
    print('\n  Horizon RMSE:')
    print(f'  {"Model":<24} {"h=1":>8} {"h=5":>8} {"h=10":>8} {"h=22":>8}')
    print('  ' + '-' * 60)
    for mname in CORE_MODELS:
        row = f'  {BASE_LABELS[mname]:<24}'
        for h in [1, 5, 10, 22]:
            val = h_rmse.get(mname, {}).get(h, np.nan)
            row += f'  {val:>8.3f}' if not np.isnan(val) else f'  {"—":>8}'
        print(row)

    # ── 8b. EOM level backtest — all 3 daily models, 2020–2025 ──
    print('\n[8b] EOM level backtest — all 3 daily models, rolling monthly 2020–2025...')
    print('     (One UC fit + 4 ARIMAX fits per month-origin — ~5 min total)')
    eom_results = month_end_eom_backtest(df, hol, start_year=2020, end_year=2025)

    print('\n[8d] EOM level backtest — pre-COVID period (2018-2019)...')
    eom_results_precovid = month_end_eom_backtest(df, hol, start_year=2018, end_year=2019)

    # ── 9. Figures ──
    print('\n[9] Generating figures...')

    # cfg_main is the primary config for all figures
    m_data = configs_results['cfg_main']

    # fig2 — Residual ACF/QQ on cfg_main training residuals (motivates GARCH)
    plot_residual_diagnostics(residuals, m_lbl)

    # fig3 — GARCH conditional volatility (drives prediction intervals)
    plot_garch_volatility(m_data['df_train'].index, old_res, garch_res)

    # fig4 — Seasonal EOM pattern by year + next-month forecast fan
    print('  Generating fig4 (seasonal pattern + model fan chart)...')
    fan_models_dict = {k: m_fitted[k] for k in
                       ('Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal')
                       if k in m_fitted}
    plot_seasonal_pattern(df, fan_models_dict, hol)

    # Model-comparison figures live in cic_forecast_v2.py (fig5_eom_level_backtest.png),
    # which evaluates every candidate on the primary EOM-level KPI.

    # ── 10. Excel ──
    print('\n[10] Exporting Excel output...')
    export_excel(df, configs_results, rolling_metrics, h_rmse, garch_res, eom_results)

    # Clean user-facing workbook
    print('  Exporting CIC_output.xlsx (Daily / Monthly EOM / Summary)...')
    export_cic_output_excel(df, configs_results, hol)

    # ── Final summary ──
    print('\n' + sep)
    print('  FINAL RESULTS')
    print(sep)
    paper_rmse = 4.96

    for cfg_key, cfg_data in configs_results.items():
        lbl      = cfg_data['train_label']
        ev       = cfg_data['eval_label']
        bm       = cfg_data['bench_metrics']
        old_rmse = bm['Daily_Baseline']['RMSE']
        best     = min(bm, key=lambda k: bm[k]['RMSE'])
        print(f'\n  ── Config ({lbl}), eval {ev} ──')
        print(f'  {"Model":<36} {"RMSE":>6}  {"vs Daily_Baseline":>11}  {"vs BOT paper":>13}')
        print('  ' + '-' * 70)
        for mname in ALL_MODELS_WITH_SS:
            if mname not in bm:
                continue
            r      = bm[mname]['RMSE']
            tag    = ' ← best' if mname == best else ''
            d_old  = r - old_rmse
            d_pap  = r - paper_rmse if cfg_key == 'cfg_benchmark' else float('nan')
            s_old  = '+' if d_old >= 0 else ''
            s_pap  = '+' if not np.isnan(d_pap) and d_pap >= 0 else ''
            d_pap_str = f'{s_pap}{d_pap:+.3f}' if not np.isnan(d_pap) else '      —'
            lbl2   = model_label(mname, lbl)
            print(f'  {lbl2:<36} {r:>6.3f}  {s_old}{d_old:>10.3f}  {d_pap_str:>13}{tag}')
        if cfg_key == 'cfg_benchmark':
            print(f'  {"[BOT 2022 paper (2017-2021)]":<36} {"4.960":>6}  {"baseline":>11}  {"0.000":>13}')

    # EOM level RMSE summary
    print(f'\n  ── EOM Level RMSE (primary KPI — 1-month-ahead, 2020–2025) ──')
    eom_models = ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']
    print(f'  {"Model":<20}  {"2024–25 RMSE":>14}  {"Overall RMSE":>14}')
    print('  ' + '-' * 52)
    old24 = None
    for k in eom_models:
        r = eom_results.get(k, {})
        if not len(r.get('dates', [])):
            continue
        mask24   = r['dates'].year >= 2024
        rmse24   = np.sqrt(np.mean(r['errors'][mask24] ** 2)) if mask24.sum() > 0 else np.nan
        rmse_all = r['RMSE']
        if k == 'Daily_Baseline':
            old24 = rmse24
        tag = ''
        if k != 'Daily_Baseline' and old24 is not None and not np.isnan(rmse24):
            tag = '  ← better' if rmse24 < old24 else ''
        print(f'  {k:<20}  {rmse24:>14.3f}  {rmse_all:>14.3f}{tag}')

    # Pre-COVID EOM RMSE
    print(f'\n  ── Pre-COVID EOM Level RMSE (2018–2019) ──')
    print(f'  {"Model":<20}  {"Overall RMSE":>14}')
    print('  ' + '-' * 38)
    for k in eom_models:
        r = eom_results_precovid.get(k, {})
        if not len(r.get('dates', [])):
            continue
        print(f'  {k:<20}  {r["RMSE"]:>14.3f}')

    # ── NEW MODEL SUMMARY ──
    print(f'\n  ── NEW MODEL SUMMARY ──')
    _pre_rmse = {}
    for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']:
        r = eom_results_precovid.get(k, {})
        _pre_rmse[k] = r.get('RMSE', float('nan'))
    print(f'  Pre-COVID OOS EOM RMSE (2018-2019):  ' +
          '  '.join(f'{k}={_pre_rmse[k]:.1f}' for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']))

    _2025_rmse = {}
    for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']:
        r = eom_results.get(k, {})
        if len(r.get('dates', [])):
            mask25 = r['dates'].year == 2025
            _2025_rmse[k] = np.sqrt(np.mean(r['errors'][mask25]**2)) if mask25.sum() > 0 else float('nan')
        else:
            _2025_rmse[k] = float('nan')
    print(f'  2025 EOM RMSE:                        ' +
          '  '.join(f'{k}={_2025_rmse[k]:.1f}' for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']))

    _comp_st = eom_results.get('_comp', {})
    _c1_rmse = {}
    _c2_rmse = {}
    for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']:
        c = _comp_st.get(k, {})
        if len(c.get('comp1', [])) and len(c.get('actual', [])):
            act  = np.array(c['actual'])
            err1 = act - np.array(c['comp1'])
            err2 = act - np.array(c['comp2'])
            _c1_rmse[k] = float(np.sqrt(np.mean(err1**2)))
            _c2_rmse[k] = float(np.sqrt(np.mean(err2**2)))
        else:
            _c1_rmse[k] = _c2_rmse[k] = float('nan')
    print(f'  Comp1 (Calendar-only) RMSE:           ' +
          '  '.join(f'{k}={_c1_rmse[k]:.1f}' for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']))
    print(f'  Comp2 (Drift-only)    RMSE:           ' +
          '  '.join(f'{k}={_c2_rmse[k]:.1f}' for k in ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal']))

    print(f'\n  All figures and cic_forecast_output.xlsx saved to: {os.path.abspath(".")}')
    print(sep + '\n')


# ═════════════════════════════════════════════════════════════════════════════
# PART 2 — v2 EOM-LEVEL KPI CANDIDATE HARNESS
# ═════════════════════════════════════════════════════════════════════════════
# Everything below evaluates candidates on the primary KPI (1-month-ahead EOM
# CIC level RMSE): the extra monthly/blended models, the expanding-window
# backtest with a selection/holdout split, the DM significance test, and the
# daily-RMSE no-regression guardrail. Entered via `--eom` (see main() below).
# Reuses the data loaders and daily models defined in Part 1 above.

CACHE_DIR   = 'backtest_cache'
COVID_START = pd.Timestamp('2020-03-01')
COVID_END   = pd.Timestamp('2020-12-31')

SELECTION_END  = pd.Timestamp('2023-12-31')   # last target month of selection window
FIRST_ORIGIN   = '2017-12-31'
LAST_ORIGIN    = '2026-03-31'                 # targets April 2026 (last complete month)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — MODELS
# ─────────────────────────────────────────────────────────────────────────────
# Every model implements:
#   fit_forecast(df_train, X_fut_df, last_level, target_month) ->
#       {'eom_fc': float, 'daily_fc': np.ndarray | None}
# df_train      : full history up to the origin (expanding window)
# X_fut_df      : Daily_Baseline dummy matrix for the target month's trading days
# last_level    : CIC level at the origin (L_M)
# target_month  : pd.Period of month M+1

class BaselineModel:
    """Daily_Baseline — the 2022 EViews model replicated: OLS on 55 calendar
    dummies + ARIMA(1,0,1) on the residuals (two-step ARIMAX)."""
    key = 'Daily_Baseline'

    def fit_forecast(self, df_train, X_fut_df, last_level, target_month):
        X_tr, _ = get_X(df_train, 'Daily_Baseline')
        mdl = TwoStepARIMAX().fit(df_train['Change'].values, X_tr)
        fc  = np.asarray(mdl.forecast(X_fut_df.values), float)
        return {'eom_fc': last_level + float(fc.sum()), 'daily_fc': fc}


class AdaptiveDriftWrapper:
    """Baseline: v1 Daily_AdaptiveDrift local-level state-space (adaptive drift on ΔCIC residuals)."""
    key = 'Daily_AdaptiveDrift'

    def fit_forecast(self, df_train, X_fut_df, last_level, target_month):
        X_tr, _ = get_X(df_train, 'Daily_Baseline')
        mdl = AdaptiveDriftModel('Daily_AdaptiveDrift').fit(df_train['Change'].values, X_tr)
        fc  = np.asarray(mdl.forecast(X_fut_df.values), float)
        return {'eom_fc': last_level + float(fc.sum()), 'daily_fc': fc}


class AdaptiveSeasonalWrapper:
    """v1 Daily_AdaptiveSeasonal: trailing-window OLS betas (so seasonal
    coefficients adapt) + a smooth-trend UC on the residuals."""
    key = 'Daily_AdaptiveSeasonal'

    def __init__(self, trailing_months=None):
        self.trailing_months = trailing_months
        if trailing_months:
            self.key = f'Daily_AdaptiveSeasonal_{trailing_months}m'

    def fit_forecast(self, df_train, X_fut_df, last_level, target_month):
        X_tr, _ = get_X(df_train, 'Daily_Baseline')
        mdl = AdaptiveSeasonalModel()
        if self.trailing_months:
            mdl.TRAILING_MONTHS = self.trailing_months
        mdl.fit(df_train['Change'].values, X_tr, dates=df_train.index)
        fc = np.asarray(mdl.forecast(X_fut_df.values), float)
        return {'eom_fc': last_level + float(fc.sum()), 'daily_fc': fc}


def _monthly_eom_series(df_train):
    """Month-end CIC level series: last observed level per calendar month."""
    lev = df_train['Currency'].dropna()
    eom = lev.groupby(lev.index.to_period('M')).last()
    eom.index = eom.index.to_timestamp('M')
    return eom


class MonthlySarimaModel:
    """Bet 1a: SARIMA(1,1,1)(0,1,1,12) on the monthly EOM level with
    2020-03 / 2020-04 intervention dummies (COVID cash-hoarding pulse)."""
    key = 'Monthly_SARIMA'

    def fit_forecast(self, df_train, X_fut_df, last_level, target_month):
        eom = _monthly_eom_series(df_train)
        exog = pd.DataFrame(index=eom.index)
        exog['covid_mar20'] = (eom.index.to_period('M') == pd.Period('2020-03')).astype(float)
        exog['covid_apr20'] = (eom.index.to_period('M') == pd.Period('2020-04')).astype(float)
        mdl = SARIMAX(eom, exog=exog, order=(1, 1, 1), seasonal_order=(0, 1, 1, 12),
                      enforce_stationarity=False, enforce_invertibility=False
                      ).fit(disp=False, maxiter=300)
        exog_f = pd.DataFrame({'covid_mar20': [0.0], 'covid_apr20': [0.0]})
        fc = float(mdl.forecast(steps=1, exog=exog_f).iloc[0])
        return {'eom_fc': fc, 'daily_fc': None}


class MonthlyUCModel:
    """Bet 1b: UC local-linear-trend + 12-month seasonal on the monthly EOM
    level, with COVID months (2020-03 → 2020-12) masked as missing."""
    key = 'Monthly_UC'

    def fit_forecast(self, df_train, X_fut_df, last_level, target_month):
        eom = _monthly_eom_series(df_train).astype(float)
        masked = eom.copy()
        masked[(masked.index >= COVID_START) & (masked.index <= COVID_END)] = np.nan
        mdl = UnobservedComponents(masked, level='local linear trend', seasonal=12
                                   ).fit(disp=False, method='bfgs', maxiter=300)
        fc = float(np.asarray(mdl.get_forecast(steps=1).predicted_mean)[0])
        return {'eom_fc': fc, 'daily_fc': None}


class LevelTrendModel:
    """Bet 2: smooth-trend UC on the calendar-adjusted CIC level.

    Step 1: full-sample OLS of ΔCIC on the Daily_Baseline dummy matrix (intercept
            included) → calendar mean m_t; cumulate residuals into the
            calendar-adjusted level C_t = Σ (ΔCIC_s − m_s).
    Step 2: UnobservedComponents(C_t, level=trend_spec, autoregressive=1),
            optionally masking COVID (2020-03 → 2020-12) as missing so the
            trend variances are not MLE'd off the hoarding shock.
    EOM forecast = L_M + Σ future m + (Ĉ_terminal − C_M).
    Daily path   = m_{t} + ΔĈ_{t} (for the daily-RMSE guardrail).
    """

    def __init__(self, trend='smooth trend', covid_nan=True, key=None):
        self.trend     = trend
        self.covid_nan = covid_nan
        suffix         = '' if trend == 'smooth trend' else '_LLT'
        self.key       = key or ('Daily_LevelTrend' + suffix
                                 + ('' if covid_nan else '_NoMask'))

    def fit_forecast(self, df_train, X_fut_df, last_level, target_month):
        y = df_train['Change'].values.astype(float)
        X_tr, _ = get_X(df_train, 'Daily_Baseline')
        ols = LinearRegression(fit_intercept=True).fit(X_tr, y)
        resid = y - ols.predict(X_tr)
        cum = pd.Series(np.cumsum(resid), index=df_train.index)
        endog = cum.copy()
        if self.covid_nan:
            endog[(endog.index >= COVID_START) & (endog.index <= COVID_END)] = np.nan
        mdl = UnobservedComponents(endog.values, level=self.trend, autoregressive=1
                                   ).fit(disp=False, method='bfgs', maxiter=300)
        n_fc  = len(X_fut_df)
        c_hat = np.asarray(mdl.get_forecast(steps=n_fc).predicted_mean, float)
        m_fut = ols.predict(X_fut_df.values)
        c_last = float(cum.iloc[-1])
        daily = m_fut + np.diff(np.concatenate([[c_last], c_hat]))
        eom_fc = last_level + float(m_fut.sum()) + (float(c_hat[-1]) - c_last)
        return {'eom_fc': eom_fc, 'daily_fc': daily}


# CLI name → model factory.  Model keys (the names that appear in every output)
# follow one scheme: <Frequency>_<Method>, and blends are Blend_<members>.
MODEL_FACTORY = {
    'baseline':             BaselineModel,               # Daily_Baseline
    'adaptive_drift':       AdaptiveDriftWrapper,        # Daily_AdaptiveDrift
    'adaptive_seasonal':    AdaptiveSeasonalWrapper,     # Daily_AdaptiveSeasonal
    'monthly_sarima':       MonthlySarimaModel,          # Monthly_SARIMA
    'monthly_uc':           MonthlyUCModel,              # Monthly_UC
    'level_trend':          lambda: LevelTrendModel(trend='smooth trend', covid_nan=True),
    'level_trend_llt':      lambda: LevelTrendModel(trend='local linear trend', covid_nan=True),
    'level_trend_nomask':   lambda: LevelTrendModel(trend='smooth trend', covid_nan=False),
}
DEFAULT_MODELS = list(MODEL_FACTORY)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — EOM BACKTEST HARNESS (expanding window, per-origin cache)
# ─────────────────────────────────────────────────────────────────────────────

def _cache_path(model_key):
    os.makedirs(CACHE_DIR, exist_ok=True)
    return os.path.join(CACHE_DIR, f'{model_key}.pkl')


def _load_cache(model_key):
    p = _cache_path(model_key)
    if os.path.exists(p):
        with open(p, 'rb') as f:
            return pickle.load(f)
    return {}


def _save_cache(model_key, cache):
    with open(_cache_path(model_key), 'wb') as f:
        pickle.dump(cache, f)


def eom_backtest(df, hol, model_names, first_origin=FIRST_ORIGIN,
                 last_origin=LAST_ORIGIN, use_cache=True, verbose=True):
    """
    For each origin = calendar month-end M with data available:
      1. train = df up to origin (expanding window), L_M = last observed level
      2. build the Daily_Baseline dummy matrix for M+1's trading days
         (generate_future_exog — same call as v1's month_end_eom_backtest)
      3. each model forecasts the EOM level of M+1 (daily models: L_M + Σ ΔCIĈ)
      4. error = actual EOM(M+1) − forecast
    Returns {model_key: DataFrame[target, actual, forecast, error, ...]} plus
    a parallel dict of daily paths for the daily-RMSE guardrail.
    """
    origins = pd.date_range(first_origin, last_origin, freq='ME')
    results = {}
    daily_store = {}

    for name in model_names:
        factory = MODEL_FACTORY[name]
        model = factory()
        cache = _load_cache(model.key) if use_cache else {}
        rows, dirty = [], False
        daily_store[model.key] = {}

        for origin in origins:
            avail = df.index[df.index <= origin]
            if len(avail) < 500:
                continue
            df_train = df.loc[:origin]

            nm_start = origin + pd.offsets.MonthBegin(1)
            nm_end   = nm_start + pd.offsets.MonthEnd(0)
            df_next  = df.loc[nm_start:nm_end]
            if len(df_next) < 5:
                continue
            lev_next = df_next['Currency'].dropna()
            if len(lev_next) == 0:
                continue
            actual_eom = float(lev_next.iloc[-1])
            lev_hist = df_train['Currency'].dropna()
            if len(lev_hist) == 0:
                continue
            last_level = float(lev_hist.iloc[-1])

            okey = origin.strftime('%Y-%m-%d')
            if okey in cache:
                out = cache[okey]
            else:
                fc_start = df_next.index[0].strftime('%Y-%m-%d')
                fc_end   = df_next.index[-1].strftime('%Y-%m-%d')
                try:
                    X_fut_df = generate_future_exog('Daily_Baseline', fc_start, fc_end, hol,
                                                    actual_dates=df_next.index)
                    if len(X_fut_df) == 0:
                        continue
                    out = model.fit_forecast(df_train, X_fut_df, last_level,
                                             nm_start.to_period('M'))
                    out['daily_dates'] = X_fut_df.index
                except Exception as exc:
                    print(f'    ⚠ {model.key} @ {okey}: {exc}')
                    continue
                cache[okey] = out
                dirty = True
                if verbose:
                    print(f'    {model.key:<24} {okey}  EOM fc={out["eom_fc"]:.1f} '
                          f'actual={actual_eom:.1f}', flush=True)

            rows.append({'origin': origin, 'target': nm_end,
                         'actual': actual_eom, 'forecast': out['eom_fc'],
                         'last_level': last_level})
            if out.get('daily_fc') is not None:
                daily_store[model.key][nm_end] = pd.Series(
                    np.asarray(out['daily_fc'], float), index=out['daily_dates'])

        if dirty and use_cache:
            _save_cache(model.key, cache)

        res = pd.DataFrame(rows).set_index('target')
        res['error'] = res['actual'] - res['forecast']
        results[model.key] = res
        if verbose:
            print(f'  {model.key:<24} n={len(res)}  overall EOM RMSE='
                  f'{np.sqrt((res["error"] ** 2).mean()):.3f}', flush=True)

    return results, daily_store


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — BET 3: FORECAST COMBINATION (post-processing on cached results)
# ─────────────────────────────────────────────────────────────────────────────

def combine_forecasts(results, member_keys, key=None, window=12, floor=0.1,
                      min_history=6):
    """
    Inverse-MSE combination of EOM forecasts, leakage-free:
    weights at origin M use only errors of targets < M's target month, over the
    trailing `window` months; floored at `floor` and renormalised.
    Before `min_history` common errors exist, fall back to equal weights.
    """
    key = key or 'Blend_' + '+'.join(member_keys)
    common = None
    for k in member_keys:
        idx = results[k].index
        common = idx if common is None else common.intersection(idx)
    common = common.sort_values()

    rows = []
    for t in common:
        fcs, ws = [], []
        for k in member_keys:
            r = results[k]
            past = r.loc[r.index < t, 'error'].tail(window)
            fcs.append(r.loc[t, 'forecast'])
            if len(past) >= min_history:
                ws.append(1.0 / max(float((past ** 2).mean()), 1e-9))
            else:
                ws.append(np.nan)
        ws = np.array(ws, float)
        if np.isnan(ws).any():
            ws = np.ones(len(member_keys))
        ws = ws / ws.sum()
        ws = np.maximum(ws, floor)
        ws = ws / ws.sum()
        fc = float(np.dot(ws, fcs))
        r0 = results[member_keys[0]].loc[t]
        rows.append({'target': t, 'origin': r0['origin'], 'actual': r0['actual'],
                     'forecast': fc, 'last_level': r0['last_level'],
                     **{f'w_{k}': w for k, w in zip(member_keys, ws)}})
    res = pd.DataFrame(rows).set_index('target')
    res['error'] = res['actual'] - res['forecast']
    return key, res


def equal_weight(results, member_keys, key=None):
    key = key or 'Blend_Avg_' + '+'.join(member_keys)
    common = None
    for k in member_keys:
        idx = results[k].index
        common = idx if common is None else common.intersection(idx)
    common = common.sort_values()
    fc = sum(results[k].loc[common, 'forecast'] for k in member_keys) / len(member_keys)
    r0 = results[member_keys[0]].loc[common]
    res = pd.DataFrame({'origin': r0['origin'], 'actual': r0['actual'],
                        'forecast': fc, 'last_level': r0['last_level']})
    res['error'] = res['actual'] - res['forecast']
    return key, res


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — METRICS, DM TEST, GUARDRAIL
# ─────────────────────────────────────────────────────────────────────────────

def rmse(e):
    e = np.asarray(e, float)
    return float(np.sqrt(np.mean(e ** 2))) if len(e) else np.nan


def summarize(results, lo=None, hi=None, label=''):
    """Per-model overall + per-year EOM RMSE over targets in [lo, hi]."""
    rows = []
    for k, r in results.items():
        m = r
        if lo is not None:
            m = m[m.index >= lo]
        if hi is not None:
            m = m[m.index <= hi]
        if len(m) == 0:
            continue
        row = {'Model': k, 'n': len(m), 'RMSE': rmse(m['error']),
               'MAE': float(m['error'].abs().mean()),
               'Bias': float(m['error'].mean())}
        for yr in sorted(m.index.year.unique()):
            row[f'RMSE_{yr}'] = rmse(m.loc[m.index.year == yr, 'error'])
        rows.append(row)
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).set_index('Model').sort_values('RMSE')
    if label:
        print(f'\n=== {label} ===')
        print(out.round(3).to_string())
    return out


def dm_test(e1, e2):
    """Diebold–Mariano test on squared errors (h=1), Harvey small-sample corr.
    Negative stat → model 1 more accurate. Returns (stat, p_value)."""
    d = np.asarray(e1, float) ** 2 - np.asarray(e2, float) ** 2
    n = len(d)
    dbar = d.mean()
    var = d.var(ddof=1) / n
    if var <= 0:
        return np.nan, np.nan
    dm = dbar / np.sqrt(var)
    harvey = np.sqrt((n + 1 - 2 * 1 + 1 * (1 - 1) / n) / n)  # h=1
    stat = dm * harvey
    p = 2 * (1 - stats.t.cdf(abs(stat), df=n - 1))
    return float(stat), float(p)


def daily_guardrail(df, daily_store, baseline_key='Daily_Baseline', lo=None, hi=None):
    """Per-year daily ΔCIC RMSE per model (only models with daily paths)."""
    rows = {}
    for k, paths in daily_store.items():
        if not paths:
            continue
        err = []
        for t, fc in paths.items():
            if lo is not None and t < lo:
                continue
            if hi is not None and t > hi:
                continue
            act = df['Change'].reindex(fc.index)
            e = (act - fc).dropna()
            err.append(e)
        if not err:
            continue
        e = pd.concat(err).sort_index()
        rows[k] = {f'{yr}': rmse(e[e.index.year == yr]) for yr in sorted(e.index.year.unique())}
        rows[k]['Overall'] = rmse(e)
    out = pd.DataFrame(rows).T
    print('\n=== Daily ΔCIC RMSE guardrail (per target-month days) ===')
    print(out.round(3).to_string())
    if baseline_key in out.index:
        base = out.loc[baseline_key]
        viol = out[out.gt(base * 1.05).any(axis=1)].index.difference([baseline_key])
        if len(viol):
            print(f'  ⚠ guardrail exceeded (+5% vs {baseline_key} in ≥1 year): {list(viol)}')
        else:
            print('  ✓ all models within +5% of baseline in every year')
    return out


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 — OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def reconciled_daily_path(daily_store, base_key, base_eom, target_eom, target):
    """Production daily path for an EOM-only forecast (M1 or a combination):
    take the daily-model's within-month calendar shape and shift it by the
    constant per-day amount needed to hit the reconciled EOM level."""
    base = daily_store[base_key][target]
    shift = (target_eom - base_eom) / len(base)
    return base + shift


def export_results(results, sel_tbl, hold_tbl, guard_tbl, path='CIC_output.xlsx'):
    """Writes EOM_Selection / EOM_Holdout / EOM_Detail / Daily_Guardrail into the
    single consolidated workbook (see _excel_writer) — append-safe, so this never
    wipes the Daily / Monthly EOM / Summary sheets from the daily pipeline."""
    with _excel_writer(path) as xw:
        sel_tbl.round(3).to_excel(xw, sheet_name='EOM_Selection')
        hold_tbl.round(3).to_excel(xw, sheet_name='EOM_Holdout')
        detail = []
        for k, r in results.items():
            d = r.reset_index()
            d.insert(0, 'Model', k)
            detail.append(d)
        pd.concat(detail).round(3).to_excel(xw, sheet_name='EOM_Detail', index=False)
        if guard_tbl is not None:
            guard_tbl.round(3).to_excel(xw, sheet_name='Daily_Guardrail')
    print(f'\n  Results exported → {path}  (sheets: EOM_Selection, EOM_Holdout, EOM_Detail, Daily_Guardrail)')


def plot_eom(results, keys, path='fig5_eom_level_backtest.png'):
    fig, axes = plt.subplots(2, 1, figsize=(13, 9), sharex=True)
    r0 = results[keys[0]]
    axes[0].plot(r0.index, r0['actual'], 'k-', lw=2, label='Actual EOM level')
    for k in keys:
        r = results[k]
        axes[0].plot(r.index, r['forecast'], lw=1.2, alpha=0.85, label=k)
        axes[1].plot(r.index, r['error'], lw=1.2, alpha=0.85, label=k)
    axes[1].axhline(0, color='k', lw=0.8)
    axes[1].axvline(SELECTION_END, color='grey', ls='--', lw=1)
    axes[1].text(SELECTION_END, axes[1].get_ylim()[1] * 0.9, ' holdout →',
                 color='grey', fontsize=9)
    axes[0].set_title('1-month-ahead end-of-month CIC level — actual vs forecast')
    axes[1].set_title('EOM forecast error (THB bn)')
    for ax in axes:
        ax.legend(fontsize=8, ncol=3)
        ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f'  Figure saved → {path}')


# Fixed categorical order (colorblind-validated: node scripts/validate_palette.js
# from the dataviz skill, slots 1/2/3/4, light mode) — identity, not ranked by
# performance, so a model keeps its color across runs/figures.
_KPI_COLOR_ORDER = ['Daily_Baseline', 'Daily_AdaptiveDrift', 'Daily_AdaptiveSeasonal',
                    'Monthly_SARIMA', 'Monthly_UC', 'Daily_LevelTrend',
                    'Blend_Baseline_Monthly', 'Blend_InvMSE_All3', 'Blend_EqualWeight']
_KPI_PALETTE = ['#2a78d6', '#eb6834', '#1baf7a', '#eda100',
               '#e87ba4', '#4a3aa7', '#e34948', '#008300', '#6d6d63']


def plot_kpi_comparison(sel_tbl, hold_tbl, path='fig6_eom_rmse_kpi.png'):
    """The primary-KPI chart: EOM level RMSE by model, selection vs. holdout.
    Ranks models by holdout RMSE (the number that decides the winner, §6 of
    CIC_model_document.md) and marks the best one — this is the harness's own
    answer to "which model wins", not a separate manual read of the tables.
    """
    common = [m for m in _KPI_COLOR_ORDER if m in sel_tbl.index and m in hold_tbl.index]
    common += [m for m in hold_tbl.index if m in sel_tbl.index and m not in common]
    if not common:
        print('  (fig6 skipped: no model has both a selection and holdout RMSE)')
        return None
    ranked = hold_tbl.loc[common, 'RMSE'].sort_values().index.tolist()
    best = ranked[0]
    colors = {m: _KPI_PALETTE[_KPI_COLOR_ORDER.index(m) % len(_KPI_PALETTE)]
             if m in _KPI_COLOR_ORDER else '#6d6d63' for m in common}

    sel_rmse  = [sel_tbl.loc[m, 'RMSE']  for m in common]
    hold_rmse = [hold_tbl.loc[m, 'RMSE'] for m in common]

    fig, ax = plt.subplots(figsize=(max(9, 1.9 * len(common) + 3), 6.5))
    x = np.arange(len(common))
    w = 0.34
    b1 = ax.bar(x - w/2, sel_rmse, width=w, color=[colors[m] for m in common],
               edgecolor='white', linewidth=0.6, label=f'Selection {SELECTION_END.year - 5}–{SELECTION_END.year} (in-sample choices)')
    b2 = ax.bar(x + w/2, hold_rmse, width=w, color=[colors[m] for m in common],
               alpha=0.55, edgecolor='white', linewidth=0.6, hatch='///',
               label=f'Holdout {SELECTION_END.year + 1}–{pd.Timestamp(LAST_ORIGIN).year} (evaluated once, decides the winner)')
    for bars in (b1, b2):
        for rect in bars:
            h = rect.get_height()
            ax.annotate(f'{h:.1f}', (rect.get_x() + rect.get_width() / 2, h),
                       xytext=(0, 3), textcoords='offset points',
                       ha='center', va='bottom', fontsize=9, color='#2b2b28')
    best_x = common.index(best)
    ax.annotate('BEST', xy=(best_x, max(sel_rmse[best_x], hold_rmse[best_x])),
               xytext=(0, 16), textcoords='offset points', ha='center',
               fontsize=9.5, fontweight='bold', color='#2b2b28',
               arrowprops=dict(arrowstyle='-', color='#2b2b28', lw=0.8))

    ax.set_xticks(x)
    ax.set_xticklabels([m.replace('_', '\n') for m in common], fontsize=9.5)
    ax.set_ylabel('EOM level RMSE (THB bn) — lower is better', fontsize=11)
    ax.set_title('Primary KPI: 1-month-ahead end-of-month CIC level RMSE\n'
                f'Winner (lowest holdout RMSE): {best} — {hold_tbl.loc[best, "RMSE"]:.1f}',
                fontsize=13, fontweight='bold')
    ax.legend(loc='upper left', fontsize=8.5, frameon=False)
    ax.grid(axis='y', alpha=0.3)
    ax.spines[['top', 'right']].set_visible(False)
    ax.set_ylim(0, max(sel_rmse + hold_rmse) * 1.25)
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'  Figure saved → {path}')
    print(f'  → BEST MODEL by holdout EOM RMSE: {best} '
         f'({hold_tbl.loc[best, "RMSE"]:.2f} vs. runner-up '
         f'{ranked[1]}={hold_tbl.loc[ranked[1], "RMSE"]:.2f})' if len(ranked) > 1 else
         f'  → BEST MODEL by holdout EOM RMSE: {best} ({hold_tbl.loc[best, "RMSE"]:.2f})')
    return best


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6 — MAIN
# ─────────────────────────────────────────────────────────────────────────────

# Reference numbers from BEFORE the 2026-08 generate_future_exog fix (§3(a)):
# the harness reproduced these exactly, which is the bug's fingerprint, not a
# target — a "phantom trading day" bias that is internally consistent between
# the old v1 and v2 harnesses (both called the same buggy function) but wrong
# against the actual RAW calendar. Kept here purely as a documented before/after.
PRE_FIX_REFERENCE = {'Daily_Baseline': 32.860, 'Daily_AdaptiveDrift': 33.041}


def gate0(df, hol, use_cache=True):
    """Post-fix sanity check over origins 2019-12-31 → 2024-12-31 (targets
    2020-01 → 2025-01): the harness must run both models without exceptions
    and produce finite, stable EOM RMSE. Also reports the swing against
    PRE_FIX_REFERENCE so the size of the bug fix stays visible."""
    print('\n──── GATE 0: post-fix sanity check (origins 2019-12 → 2024-12) ────')
    results, _ = eom_backtest(df, hol, ['baseline', 'adaptive_drift'],
                              first_origin='2019-12-31', last_origin='2024-12-31',
                              use_cache=use_cache, verbose=True)
    ok = True
    for k, pre in PRE_FIX_REFERENCE.items():
        got = rmse(results[k]['error'])
        sane = np.isfinite(got) and len(results[k]) > 40
        ok &= sane
        print(f'  {k:<22} pre-fix (buggy) reference: {pre:.3f}   post-fix: {got:.3f}   '
              f'Δ={got - pre:+.3f}   {"✓ sane" if sane else "✗ FAILED"}')
    print('  GATE 0 ' + ('PASSED' if ok else 'FAILED — investigate before trusting results'))
    return ok


def run_eom_harness(args):
    """v2 EOM-level KPI harness: candidate models, selection/holdout split,
    DM test, daily guardrail, forecast combination. Entered via --eom."""
    use_cache = not args.no_cache

    print('Loading data…', flush=True)
    df  = load_data('input.xlsx')
    hol = load_holiday('input.xlsx')
    print(f'  {len(df)} rows, {df.index.min().date()} → {df.index.max().date()}')

    if args.gate0:
        gate0(df, hol, use_cache=use_cache)
        return

    model_names = [m.strip() for m in args.models.split(',') if m.strip()]
    print(f'\n──── EOM backtest: {model_names} '
          f'(origins {args.first_origin} → {args.last_origin}) ────', flush=True)
    results, daily_store = eom_backtest(df, hol, model_names,
                                        first_origin=args.first_origin,
                                        last_origin=args.last_origin,
                                        use_cache=use_cache)

    # Bet 3 — combinations (selection window decides members; defaults below).
    # Blend_Baseline_Monthly (the recommended model) only needs its own two
    # members; the 3-way blends are extra and only run when Daily_LevelTrend
    # was actually requested (it is not in DEFAULT_MODELS' fast path).
    have = set(results)
    if {'Daily_Baseline', 'Monthly_SARIMA'} <= have:
        k, r = combine_forecasts(results, ['Daily_Baseline', 'Monthly_SARIMA'],
                                 key='Blend_Baseline_Monthly')   # ← v2 winner
        results[k] = r
    if {'Daily_Baseline', 'Monthly_SARIMA', 'Daily_LevelTrend'} <= have:
        k, r = combine_forecasts(results, ['Daily_Baseline', 'Monthly_SARIMA', 'Daily_LevelTrend'],
                                 key='Blend_InvMSE_All3')
        results[k] = r
        k, r = equal_weight(results, ['Daily_Baseline', 'Monthly_SARIMA', 'Daily_LevelTrend'],
                            key='Blend_EqualWeight')
        results[k] = r
        for pair in [('Daily_Baseline', 'Daily_LevelTrend'),
                     ('Monthly_SARIMA', 'Daily_LevelTrend')]:
            k, r = combine_forecasts(results, list(pair))
            results[k] = r

    sel_tbl  = summarize(results, hi=SELECTION_END,
                         label='SELECTION window (targets 2018-01 → 2023-12)')
    hold_tbl = summarize(results, lo=SELECTION_END + pd.Timedelta(days=1),
                         label='HOLDOUT window (targets 2024-01 → 2026-04)')

    # DM tests vs Daily_Baseline on the holdout
    if 'Daily_Baseline' in results:
        print('\n=== Diebold–Mariano vs Daily_Baseline (holdout, squared errors) ===')
        base = results['Daily_Baseline']
        base_h = base[base.index > SELECTION_END]
        for k, r in results.items():
            if k == 'Daily_Baseline':
                continue
            rh = r[r.index > SELECTION_END]
            common = base_h.index.intersection(rh.index)
            if len(common) < 8:
                continue
            stat, p = dm_test(rh.loc[common, 'error'], base_h.loc[common, 'error'])
            print(f'  {k:<24} DM={stat:+.2f}  p={p:.3f}  '
                  f'({"better" if stat < 0 else "worse"} than baseline)')

    guard_tbl = daily_guardrail(df, daily_store)

    export_results(results, sel_tbl, hold_tbl, guard_tbl)
    plot_eom(results, [k for k in results if k in
                       ('Daily_Baseline', 'Daily_AdaptiveDrift', 'Monthly_SARIMA',
                        'Daily_LevelTrend', 'Blend_InvMSE_All3',
                        'Blend_Baseline_Monthly')])
    plot_kpi_comparison(sel_tbl, hold_tbl)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN — dispatch between the full v1 diagnostic pipeline and the v2 EOM harness
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description='CIC forecasting. Default: full daily diagnostic pipeline '
                    '(fig1-4, GARCH, rolling/horizon backtests, CIC_output.xlsx). '
                    'Pass --eom to instead run the v2 EOM-level KPI harness '
                    '(candidates, selection/holdout, blends, fig5).')
    ap.add_argument('--eom', action='store_true',
                    help='run the EOM-level backtest harness instead of the full pipeline')
    ap.add_argument('--gate0', action='store_true',
                    help='EOM harness only: post-fix reproduction/sanity check, then exit')
    ap.add_argument('--models', default=','.join(DEFAULT_MODELS),
                    help='EOM harness only: comma list from: ' + ','.join(MODEL_FACTORY))
    ap.add_argument('--no-cache', action='store_true')
    ap.add_argument('--first-origin', default=FIRST_ORIGIN)
    ap.add_argument('--last-origin', default=LAST_ORIGIN)
    args = ap.parse_args()

    if args.gate0 or args.eom:
        run_eom_harness(args)
    else:
        run_full_pipeline()


if __name__ == '__main__':
    main()
